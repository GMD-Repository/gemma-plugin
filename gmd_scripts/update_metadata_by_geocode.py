# ***************************************************************************
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU General Public License as published by  *
# *   the Free Software Foundation; either version 2 of the License, or     *
# *   (at your option) any later version.                                   *
# *                                                                         *
# ***************************************************************************

import os
import openpyxl
from typing import Any, Optional

from PyQt5.QtCore import QVariant
from qgis.core import (
    NULL,
    QgsField,
    QgsFields,
    QgsFeature,
    QgsFeatureSink,
    QgsProcessing,
    QgsProcessingAlgorithm,
    QgsProcessingContext,
    QgsProcessingException,
    QgsProcessingFeedback,
    QgsProcessingParameterFeatureSink,
    QgsProcessingParameterFeatureSource,
    QgsProcessingParameterField,
    QgsProcessingParameterString,
    QgsProcessingParameterFile,
    QgsProcessingParameterBoolean,
    QgsCoordinateReferenceSystem,
    QgsCoordinateTransform,
    QgsProject,
    QgsVectorLayer,
)
from PyQt5.QtGui import QIcon


def normalize_geocode(val: Any) -> str:
    """Standardizes geocode strings by stripping floating point formats, spaces, and leading zero issues."""
    if val is None or val == NULL:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = "".join(c for c in s if c.isdigit())
    if len(s) in (8, 9):
        s = s.zfill(9)
    elif len(s) in (9, 10):
        s = s.zfill(10)
    return s


def get_unique_filepath(directory: str, base_name: str, ext: str = ".gpkg") -> str:
    """Generate a unique file path by appending (1), (2), etc. if the file exists and is locked."""
    candidate = os.path.normpath(os.path.join(directory, f"{base_name}{ext}")).replace("\\", "/")
    if not os.path.exists(candidate):
        return candidate

    # Attempt to remove if existing and not locked by QGIS or another process
    try:
        os.remove(candidate)
        return candidate
    except Exception:
        pass

    # Browser-style auto-numbering (1), (2), (3)...
    counter = 1
    while True:
        candidate = os.path.normpath(os.path.join(directory, f"{base_name} ({counter}){ext}")).replace("\\", "/")
        if not os.path.exists(candidate):
            return candidate
        try:
            os.remove(candidate)
            return candidate
        except Exception:
            counter += 1


class UpdateLguByGeocodeAlgorithm(QgsProcessingAlgorithm):
    """
    Processing algorithm to update LGU boundary layers using PSGC geocode left-join.
    """

    INPUT = "INPUT"
    GEOCODE_FIELD = "GEOCODE_FIELD"
    BARANGAY_FIELD = "BARANGAY_FIELD"
    SOURCE = "SOURCE"
    SOURCE_YEAR = "SOURCE_YEAR"
    OUTPUT_DIR = "OUTPUT_DIR"
    OPEN_OUTPUT = "OPEN_OUTPUT"
    OUTPUT = "OUTPUT"

    def name(self) -> str:
        return "update_lgu_by_geocode"

    def displayName(self) -> str:
        return "Update Metadata (by Geocode)"

    def group(self) -> str:
        return "GMD Toolkits"

    def groupId(self) -> str:
        return "gmdtoolkits"

    def icon(self):
        icon_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'icons', 'update.svg')
        if os.path.exists(icon_path):
            return QIcon(icon_path)
        return QIcon(":/images/themes/default/mActionFilter.svg")

    def shortHelpString(self) -> str:
        return (
            "This algorithm updates an input LGU boundary polygon layer by performing a left-join "
            "with a PSGC Excel file using the LGU Geocode field.\n\n"
            "Parameters:\n"
            "- LGU boundary layer (polygon): Spatial polygon layer to enrich.\n"
            "- LGU geocode field: The field in the LGU layer containing geocodes.\n"
            "- LGU barangay name field: The field in the LGU layer containing original barangay names.\n"
            "- PSGC File: Excel file containing PSGC metadata (defaults to references/PSGC Q4.xlsx).\n"
            "- Source: Metadata source text (defaults to 'LGU').\n"
            "- Source Year: Metadata source year (defaults to '2026').\n"
            "- Output Directory: Folder path to save the generated GeoPackage file (defaults to location of LGU layer).\n"
            "- Show output file after running: Load the resulting layer into QGIS upon completion.\n"
        )

    def initAlgorithm(self, config: Optional[dict[str, Any]] = None):
        # Default PSGC file path
        default_psgc = os.path.abspath(
            os.path.join(os.path.dirname(os.path.dirname(__file__)), "references", "PSGC Q4.xlsx")
        )

        # 1. LGU boundary layer (polygon)
        self.addParameter(
            QgsProcessingParameterFeatureSource(
                self.INPUT,
                "LGU boundary layer (polygon)",
                [QgsProcessing.SourceType.TypeVectorPolygon],
            )
        )

        # 2. LGU geocode field (dropdown based on LGU boundary layer)
        self.addParameter(
            QgsProcessingParameterField(
                self.GEOCODE_FIELD,
                "LGU geocode field",
                parentLayerParameterName=self.INPUT,
                type=QgsProcessingParameterField.Any,
                defaultValue="geocode",
                optional=True,
            )
        )

        # 3. LGU barangay name field (dropdown based on LGU boundary layer)
        self.addParameter(
            QgsProcessingParameterField(
                self.BARANGAY_FIELD,
                "LGU barangay name field",
                parentLayerParameterName=self.INPUT,
                type=QgsProcessingParameterField.Any,
                defaultValue="barangay",
                optional=True,
            )
        )

        # 4. Source (text box, prefilled with 'LGU')
        self.addParameter(
            QgsProcessingParameterString(
                self.SOURCE,
                "Source",
                defaultValue="LGU",
            )
        )

        # 6. Source Year (text box, prefilled with 2026)
        self.addParameter(
            QgsProcessingParameterString(
                self.SOURCE_YEAR,
                "Source Year",
                defaultValue="2026",
            )
        )

        # 7. Output Directory (folder path)
        self.addParameter(
            QgsProcessingParameterFile(
                self.OUTPUT_DIR,
                "Output Directory",
                behavior=QgsProcessingParameterFile.Folder,
                defaultValue="",
                optional=True,
            )
        )

        # 8. Show output file after running
        self.addParameter(
            QgsProcessingParameterBoolean(
                self.OPEN_OUTPUT,
                "Show output file after running",
                defaultValue=True,
            )
        )

        # 9. Output Feature Sink
        self.addParameter(
            QgsProcessingParameterFeatureSink(self.OUTPUT, "Updated LGU Layer (EPSG:4326)")
        )

    def processAlgorithm(
        self,
        parameters: dict[str, Any],
        context: QgsProcessingContext,
        feedback: QgsProcessingFeedback,
    ) -> dict[str, Any]:

        source = self.parameterAsSource(parameters, self.INPUT, context)
        if source is None:
            raise QgsProcessingException(self.invalidSourceError(parameters, self.INPUT))

        geocode_field_name = self.parameterAsString(parameters, self.GEOCODE_FIELD, context)
        barangay_field_name = self.parameterAsString(parameters, self.BARANGAY_FIELD, context)
        source_val = self.parameterAsString(parameters, self.SOURCE, context)
        source_year_val = self.parameterAsString(parameters, self.SOURCE_YEAR, context)
        output_dir = self.parameterAsString(parameters, self.OUTPUT_DIR, context).strip()
        open_output = self.parameterAsBoolean(parameters, self.OPEN_OUTPUT, context)

        field_names = [f.name() for f in source.fields()]

        # Intelligent candidate fallback for geocode field if not explicitly selected
        if not geocode_field_name or geocode_field_name not in field_names:
            for candidate in ["geocode", "psgc", "code", "lgu_code", "psgc_code", "geo_code"]:
                for f in field_names:
                    if candidate in f.lower().replace("_", ""):
                        geocode_field_name = f
                        break
                if geocode_field_name and geocode_field_name in field_names:
                    break

        if not geocode_field_name or geocode_field_name not in field_names:
            raise QgsProcessingException(
                "Could not auto-detect or find the LGU geocode field in the selected layer. "
                "Please select the LGU geocode field manually."
            )

        # Intelligent candidate fallback for barangay field if not explicitly selected
        if not barangay_field_name or barangay_field_name not in field_names:
            for candidate in ["barangay", "lgu_bgy_name", "bgy_name", "brgy_name", "brgy", "bgy"]:
                for f in field_names:
                    if candidate in f.lower().replace("_", ""):
                        barangay_field_name = f
                        break
                if barangay_field_name and barangay_field_name in field_names:
                    break

        if not barangay_field_name or barangay_field_name not in field_names:
            barangay_field_name = geocode_field_name

        feedback.pushInfo(f"Using Geocode Field: '{geocode_field_name}'")
        feedback.pushInfo(f"Using Barangay Name Field: '{barangay_field_name}'")

        # Built-in PSGC file path from references directory
        psgc_file_path = os.path.abspath(
            os.path.join(os.path.dirname(os.path.dirname(__file__)), "references", "PSGC Q4.xlsx")
        )
        if not os.path.exists(psgc_file_path):
            rel_path = os.path.abspath(os.path.join(os.getcwd(), "references", "PSGC Q4.xlsx"))
            if os.path.exists(rel_path):
                psgc_file_path = rel_path
            else:
                raise QgsProcessingException(f"Built-in PSGC File not found at: '{psgc_file_path}'")

        # Read PSGC reference records (using native QgsVectorLayer OGR reader with pure Python fallback)
        feedback.pushInfo(f"Reading PSGC File: {psgc_file_path}...")
        psgc_lookup = {}
        row_count = 0
        target_cols = [
            "map_uuid", "geocode", "region", "province", "city_mun", "barangay",
            "province_code", "city_mun_code", "barangay_code", "hhcount", "bldgcount"
        ]

        def _clean_segment(v, length=0):
            if v is None or v == NULL:
                return ""
            s = str(v).strip()
            if s.endswith(".0"):
                s = s[:-2]
            s = "".join(c for c in s if c.isdigit())
            if s and length > 0:
                s = s.zfill(length)
            return s

        def _parse_int(v):
            if v is None or v == NULL:
                return None
            s = str(v).strip()
            if s == "" or s.lower() in ("nan", "none", "null"):
                return None
            try:
                return int(float(s))
            except (ValueError, TypeError):
                return None

        # Method 1: Try QgsVectorLayer via GDAL/OGR (Thread-safe & Fast)
        psgc_layer = QgsVectorLayer(f"{psgc_file_path}|layername=PSGC", "psgc_ref", "ogr")
        if not psgc_layer or not psgc_layer.isValid():
            psgc_layer = QgsVectorLayer(psgc_file_path, "psgc_ref", "ogr")

        if psgc_layer and psgc_layer.isValid():
            fields = psgc_layer.fields()
            col_indices = {}
            for col_name in target_cols:
                col_norm = col_name.replace("_", "").lower()
                for field in fields:
                    if field.name().lower().replace("_", "").replace("/", "") == col_norm:
                        col_indices[col_name] = field.name()
                        break

            if "geocode" in col_indices or "province_code" in col_indices:
                for feat in psgc_layer.getFeatures():
                    if feedback.isCanceled():
                        break

                    raw_geo = feat.attribute(col_indices["geocode"]) if "geocode" in col_indices else ""
                    norm_geo = normalize_geocode(raw_geo)

                    prov_c = _clean_segment(feat.attribute(col_indices["province_code"])) if "province_code" in col_indices else ""
                    city_c = _clean_segment(feat.attribute(col_indices["city_mun_code"]), 2) if "city_mun_code" in col_indices else ""
                    bgy_c = _clean_segment(feat.attribute(col_indices["barangay_code"]), 3) if "barangay_code" in col_indices else ""

                    concat_key = f"{prov_c}{city_c}{bgy_c}" if (prov_c and city_c and bgy_c) else ""
                    
                    def _get_val(key):
                        if key in col_indices:
                            v = feat.attribute(col_indices[key])
                            return str(v).strip() if v is not None and v != NULL else ""
                        return ""

                    rec = {
                        "map_uuid": _get_val("map_uuid"),
                        "geocode": norm_geo if norm_geo else concat_key,
                        "region": _get_val("region"),
                        "province": _get_val("province"),
                        "city_mun": _get_val("city_mun"),
                        "barangay": _get_val("barangay"),
                        "hhcount": _parse_int(feat.attribute(col_indices["hhcount"])) if "hhcount" in col_indices else None,
                        "bldgcount": _parse_int(feat.attribute(col_indices["bldgcount"])) if "bldgcount" in col_indices else None,
                    }

                    # Index by multiple keys for maximum match robustness
                    if concat_key:
                        psgc_lookup[concat_key] = rec
                        psgc_lookup[concat_key.lstrip("0")] = rec
                    if norm_geo:
                        psgc_lookup[norm_geo] = rec
                        if len(norm_geo) >= 8:
                            psgc_lookup[norm_geo[:8]] = rec
                            if norm_geo.startswith("13") or norm_geo.startswith("0"):
                                psgc_lookup[norm_geo[2:10]] = rec

                    row_count += 1

        # Method 2: Fallback to openpyxl with read_only=False (pure Python XML parsing, thread-safe)
        if row_count == 0:
            try:
                wb = openpyxl.load_workbook(psgc_file_path, read_only=False, data_only=True)
                sheet_name = "PSGC" if "PSGC" in wb.sheetnames else wb.sheetnames[0]
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)
                header_row = next(rows_iter)
                header_clean = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]

                col_indices = {}
                for col_name in target_cols:
                    if col_name in header_clean:
                        col_indices[col_name] = header_clean.index(col_name)

                if "geocode" in col_indices or "province_code" in col_indices:
                    for row in rows_iter:
                        if feedback.isCanceled():
                            break
                        raw_geo = row[col_indices["geocode"]] if "geocode" in col_indices else ""
                        norm_geo = normalize_geocode(raw_geo)

                        prov_c = _clean_segment(row[col_indices["province_code"]]) if "province_code" in col_indices else ""
                        city_c = _clean_segment(row[col_indices["city_mun_code"]], 2) if "city_mun_code" in col_indices else ""
                        bgy_c = _clean_segment(row[col_indices["barangay_code"]], 3) if "barangay_code" in col_indices else ""

                        concat_key = f"{prov_c}{city_c}{bgy_c}" if (prov_c and city_c and bgy_c) else ""

                        def _row_val(k):
                            if k in col_indices and row[col_indices[k]] is not None:
                                return str(row[col_indices[k]]).strip()
                            return ""

                        rec = {
                            "map_uuid": _row_val("map_uuid"),
                            "geocode": norm_geo if norm_geo else concat_key,
                            "region": _row_val("region"),
                            "province": _row_val("province"),
                            "city_mun": _row_val("city_mun"),
                            "barangay": _row_val("barangay"),
                            "hhcount": _parse_int(row[col_indices["hhcount"]]) if "hhcount" in col_indices else None,
                            "bldgcount": _parse_int(row[col_indices["bldgcount"]]) if "bldgcount" in col_indices else None,
                        }

                        if concat_key:
                            psgc_lookup[concat_key] = rec
                            psgc_lookup[concat_key.lstrip("0")] = rec
                        if norm_geo:
                            psgc_lookup[norm_geo] = rec
                            if len(norm_geo) >= 8:
                                psgc_lookup[norm_geo[:8]] = rec

                        row_count += 1
                wb.close()
            except Exception as e:
                raise QgsProcessingException(f"Failed to read PSGC file: {e}")

        feedback.pushInfo(f"Loaded {row_count} PSGC reference records into lookup table.")

        def _field(name, ftype, length=0):
            f = QgsField(name, ftype)
            if length:
                f.setLength(length)
            return f

        output_fields = QgsFields()
        output_fields.append(_field("fid",          QVariant.Int))
        output_fields.append(_field("map_uuid",      QVariant.String, 36))
        output_fields.append(_field("geocode",       QVariant.String, 50))
        output_fields.append(_field("region",        QVariant.String, 100))
        output_fields.append(_field("province",      QVariant.String, 100))
        output_fields.append(_field("city_mun",      QVariant.String, 100))
        output_fields.append(_field("barangay",      QVariant.String, 100))
        output_fields.append(_field("code",          QVariant.String, 50))
        output_fields.append(_field("remarks",       QVariant.String, 255))
        output_fields.append(_field("source",        QVariant.String, 100))
        output_fields.append(_field("hhcount",       QVariant.Int))
        output_fields.append(_field("bldgcount",     QVariant.Int))
        output_fields.append(_field("sy",            QVariant.String, 10))
        output_fields.append(_field("boundary",      QVariant.String, 20))
        output_fields.append(_field("lgu_bgy_name",  QVariant.String, 100))
        output_fields.append(_field("bdry_status",   QVariant.String, 20))

        # CRS Reprojection Setup (WGS 84 / EPSG:4326)
        target_crs = QgsCoordinateReferenceSystem("EPSG:4326")
        reproject = source.sourceCrs().authid() != "EPSG:4326"
        transform = None
        if reproject:
            transform = QgsCoordinateTransform(source.sourceCrs(), target_crs, context.transformContext())
            feedback.pushInfo(f"Reprojecting from {source.sourceCrs().authid()} to EPSG:4326...")

        # Determine output directory (physical layer folder on disk -> project folder -> home path -> Documents)
        if not output_dir:
            input_layer = QgsProcessingUtils.mapLayerFromString(str(parameters.get(self.INPUT)), context)
            if input_layer and input_layer.isValid():
                src_path = input_layer.source().split("|")[0]
                if os.path.isfile(src_path) or os.path.isdir(src_path):
                    output_dir = os.path.dirname(src_path) if os.path.isfile(src_path) else src_path

        if not output_dir or not os.path.isdir(output_dir):
            proj_file = QgsProject.instance().fileName()
            if proj_file and os.path.isfile(proj_file):
                output_dir = os.path.dirname(proj_file)
            elif QgsProject.instance().homePath() and os.path.isdir(QgsProject.instance().homePath()):
                output_dir = QgsProject.instance().homePath()
            else:
                output_dir = os.path.expanduser("~/Documents")

        # Determine custom GeoPackage filename
        detected_prefix = ""
        for feat in source.getFeatures():
            raw_val = feat.attribute(geocode_field_name)
            norm_v = normalize_geocode(raw_val)
            if norm_v and len(norm_v) >= 5:
                detected_prefix = norm_v[:5]
                break

        custom_name = f"{detected_prefix}_bgy" if detected_prefix else "Updated_LGU_by_Geocode"

        if output_dir:
            if not os.path.exists(output_dir):
                try:
                    os.makedirs(output_dir)
                except Exception as e:
                    feedback.pushInfo(f"Warning: Could not create output dir: {e}")

            gpkg_path = get_unique_filepath(output_dir, custom_name, ".gpkg")
            final_name = os.path.splitext(os.path.basename(gpkg_path))[0]
            self.custom_name = final_name

            parameters[self.OUTPUT] = gpkg_path
            feedback.pushInfo(f"Saving GeoPackage output to: {gpkg_path}")

        # Create output sink
        (sink, dest_id) = self.parameterAsSink(
            parameters,
            self.OUTPUT,
            context,
            output_fields,
            source.wkbType(),
            target_crs,
        )
        if sink is None:
            raise QgsProcessingException(self.invalidSinkError(parameters, self.OUTPUT))

        self.dest_id = dest_id
        self.open_output = open_output

        # Process features and perform left-join
        total_feats = source.featureCount()
        step = 100.0 / total_feats if total_feats else 0
        current = 0

        def _attr_or_null(val):
            return val if (val is not None and val != "" and val != NULL) else NULL

        for feat in source.getFeatures():
            if feedback.isCanceled():
                break

            geom = feat.geometry()
            if reproject and transform and not geom.isEmpty():
                geom.transform(transform)

            raw_geo_val = feat.attribute(geocode_field_name)
            lgu_geo_str = str(raw_geo_val).strip() if raw_geo_val is not None and raw_geo_val != NULL else ""
            norm_geo_val = normalize_geocode(raw_geo_val)
            lgu_key_8 = norm_geo_val[:8] if len(norm_geo_val) >= 8 else norm_geo_val

            raw_bgy_val = feat.attribute(barangay_field_name)
            lgu_bgy_name_str = str(raw_bgy_val).strip() if raw_bgy_val is not None and raw_bgy_val != NULL else ""

            # Perform key lookup using first 8 chars, full code, or fallback keys
            psgc_match = (
                psgc_lookup.get(lgu_key_8)
                or psgc_lookup.get(lgu_key_8.lstrip("0"))
                or psgc_lookup.get(norm_geo_val)
                or (psgc_lookup.get(norm_geo_val[2:10]) if len(norm_geo_val) >= 10 else None)
                or {}
            )

            out_feat = QgsFeature(output_fields)
            out_feat.setGeometry(geom)

            fid_idx = output_fields.indexOf("fid")
            fid_val = current + 1
            if fid_idx != -1:
                out_feat.setAttribute(fid_idx, fid_val)
            out_feat.setId(fid_val)

            out_feat.setAttribute("map_uuid", _attr_or_null(psgc_match.get("map_uuid")))
            out_feat.setAttribute("geocode", _attr_or_null(lgu_geo_str if lgu_geo_str else norm_geo_val))
            out_feat.setAttribute("region", _attr_or_null(psgc_match.get("region")))
            out_feat.setAttribute("province", _attr_or_null(psgc_match.get("province")))
            out_feat.setAttribute("city_mun", _attr_or_null(psgc_match.get("city_mun")))
            out_feat.setAttribute("barangay", _attr_or_null(psgc_match.get("barangay")))
            out_feat.setAttribute("code", "1003")
            out_feat.setAttribute("remarks", NULL)
            out_feat.setAttribute("source", _attr_or_null(source_val))
            out_feat.setAttribute("hhcount", _attr_or_null(psgc_match.get("hhcount")))
            out_feat.setAttribute("bldgcount", _attr_or_null(psgc_match.get("bldgcount")))
            out_feat.setAttribute("sy", _attr_or_null(source_year_val))
            out_feat.setAttribute("boundary", "Barangay")
            out_feat.setAttribute("lgu_bgy_name", _attr_or_null(lgu_bgy_name_str))
            out_feat.setAttribute("bdry_status", NULL)

            sink.addFeature(out_feat, QgsFeatureSink.FastInsert)
            current += 1
            feedback.setProgress(int(current * step))
            feedback.setProgress(int(current * step))

        feedback.pushInfo(f"Successfully processed {current} features into updated layer.")
        return {self.OUTPUT: dest_id}

    def postProcessAlgorithm(self, context: QgsProcessingContext, feedback: QgsProcessingFeedback) -> dict[str, Any]:
        layer_name = getattr(self, "custom_name", "Updated_LGU_by_Geocode")
        dest_id = getattr(self, "dest_id", None)
        open_output = getattr(self, "open_output", True)

        if dest_id and open_output:
            if context.willLoadLayerOnCompletion(dest_id):
                details = context.layerToLoadOnCompletionDetails(dest_id)
                details.name = layer_name
            else:
                target_source = dest_id.split("|")[0]
                already_loaded = False
                for l in QgsProject.instance().mapLayers().values():
                    if l.source().split("|")[0] == target_source:
                        already_loaded = True
                        l.setName(layer_name)
                        break

                if not already_loaded:
                    new_layer = QgsVectorLayer(dest_id, layer_name, "ogr")
                    if new_layer and new_layer.isValid():
                        QgsProject.instance().addMapLayer(new_layer)
                        feedback.pushInfo(f"Loaded output layer '{layer_name}' into QGIS canvas.")

        return {self.OUTPUT: dest_id}

    def createInstance(self):
        return self.__class__()
