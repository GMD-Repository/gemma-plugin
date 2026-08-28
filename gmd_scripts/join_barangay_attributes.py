# ***************************************************************************
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU General Public License as published by  *
# *   the Free Software Foundation; either version 2 of the License, or     *
# *   (at your option) any later version.                                   *
# *                                                                         *
# ***************************************************************************

"""
Join Barangay Attributes — Enhanced Fuzzy Match

Timestamp   : 2026-08-28
Version     : 1.4.2
Changelog   :
  v1.0.0 - Initial enhanced version (Python-based fuzzy matching, Roman
           numeral <-> Arabic number normalization, match_status and
           all_candidates reporting columns, both outputs forced to
           temporary scratch layers so Batch Processing behaves like a
           single run).
  v1.1.0 - Removed "Generate temporary layer of filtered PSGC data"
           checkbox; the filtered PSGC layer is now always generated
           (mandatory, no longer optional).
         - Hid the Advanced Parameters section from the dialog by
           flagging "Max Levenshtein Distance" as FlagHidden (was
           FlagAdvanced) so it no longer appears in the UI at all;
           the value is still fixed internally at its default of 3.
         - Removed the "Unmatched Barangays List" output entirely,
           including its processing parameter, sink creation, and the
           logic that scheduled it to load into the project. Only the
           "Matched Barangays" output remains.
  v1.2.0 - Retained only the first column (input barangay field), second
           column ('barangay name (Final Name)'), and 'error_detail' column
           in matched output. Removed all other intermediate diagnostic
           and exact/fuzzy candidate columns.
  v1.3.0 - Embedded built-in PSGC reference dataset (references/PSGC Q4.xlsx)
           directly into the script.
  v1.4.0 - Full architectural refactor aligned with update_metadata_by_geocode.py.
  v1.4.1 - Restored Filtered PSGC Table feature sink output with thread-safe
           postProcessAlgorithm layer naming.
  v1.4.2 - Deduplicated layersToLoadOnCompletion and active project map layers
           in postProcessAlgorithm to ensure exactly one Filtered PSGC Table
           named <code_filter>_<city_name> (Filtered PSGC) is output.
"""

import os
import re
import openpyxl
from typing import Any, Optional, Dict, List, Tuple

from PyQt5.QtCore import QVariant
from qgis.core import (
    NULL,
    QgsField,
    QgsFields,
    QgsFeature,
    QgsFeatureSink,
    QgsProcessing,
    QgsProcessingAlgorithm,
    QgsProcessingContext,
    QgsProcessingException,
    QgsProcessingFeedback,
    QgsProcessingParameterFeatureSink,
    QgsProcessingParameterFeatureSource,
    QgsProcessingParameterField,
    QgsProcessingParameterNumber,
    QgsProcessingParameterDefinition,
    QgsCoordinateReferenceSystem,
    QgsProject,
    QgsVectorLayer,
    QgsProcessingUtils,
    QgsWkbTypes,
)
from PyQt5.QtGui import QIcon


# ─── Roman Numeral & Name Utilities ────────────────────────────────────────

ROMAN_TO_INT = {
    'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5,
    'VI': 6, 'VII': 7, 'VIII': 8, 'IX': 9, 'X': 10,
    'XI': 11, 'XII': 12, 'XIII': 13, 'XIV': 14, 'XV': 15,
    'XVI': 16, 'XVII': 17, 'XVIII': 18, 'XIX': 19, 'XX': 20,
    'XXI': 21, 'XXII': 22, 'XXIII': 23, 'XXIV': 24, 'XXV': 25,
    'XXVI': 26, 'XXVII': 27, 'XXVIII': 28, 'XXIX': 29, 'XXX': 30,
    'XXXI': 31, 'XXXII': 32, 'XXXIII': 33, 'XXXIV': 34, 'XXXV': 35,
    'XXXVI': 36, 'XXXVII': 37, 'XXXVIII': 38, 'XXXIX': 39, 'XL': 40,
    'XLI': 41, 'XLII': 42, 'XLIII': 43, 'XLIV': 44, 'XLV': 45,
    'XLVI': 46, 'XLVII': 47, 'XLVIII': 48, 'XLIX': 49, 'L': 50,
}
INT_TO_ROMAN = {v: k for k, v in ROMAN_TO_INT.items()}

ROMAN_PATTERN = re.compile(
    r'\b(L|XL(?:IX|IV|V?I{0,3})|XXX(?:IX|IV|V?I{0,3})|XX(?:IX|IV|V?I{0,3})|X(?:IX|IV|V?I{0,3})|IX|IV|V?I{1,3})\b'
)

ABBREVIATIONS = {
    'brgy.': 'barangay', 'brgy': 'barangay',
    'sto.': 'santo', 'sta.': 'santa',
    'sr.': 'senior', 'jr.': 'junior',
    'pob.': 'poblacion',
    'mt.': 'mount', 'st.': 'saint',
}


def normalize_name(name: Any) -> str:
    """Normalize barangay name by converting to lower case, expanding abbreviations, and removing dashes."""
    if name is None or name == NULL or str(name).strip() == '' or str(name) == 'NULL':
        return ''
    text = str(name).strip().lower()
    for abbr, full in ABBREVIATIONS.items():
        text = text.replace(abbr, full)
    text = re.sub(r'[-–—]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def roman_to_arabic(name: str) -> str:
    """Convert Roman numerals in string to Arabic numbers."""
    if not name:
        return ''

    def replace_roman(match):
        roman = match.group(1).upper()
        if roman in ROMAN_TO_INT:
            return str(ROMAN_TO_INT[roman])
        return match.group(0)

    return ROMAN_PATTERN.sub(replace_roman, str(name).upper()).lower()


def arabic_to_roman(name: str) -> str:
    """Convert Arabic numbers in string to Roman numerals."""
    if not name:
        return ''

    def replace_arabic(match):
        num = int(match.group(0))
        if num in INT_TO_ROMAN:
            return INT_TO_ROMAN[num]
        return match.group(0)

    return re.sub(r'\b(\d{1,2})\b', replace_arabic, str(name)).lower()


def levenshtein_distance(s1: str, s2: str) -> int:
    """Compute exact Levenshtein edit distance between two strings."""
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)
    if len(s2) == 0:
        return len(s1)

    prev_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        curr_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = prev_row[j + 1] + 1
            deletions = curr_row[j] + 1
            substitutions = prev_row[j] + (c1 != c2)
            curr_row.append(min(insertions, deletions, substitutions))
        prev_row = curr_row

    return prev_row[-1]


def _split_name_and_number(name: str) -> Tuple[str, Optional[str]]:
    match = re.match(r'^(.+?)\s+(\d+)$', name.strip())
    if match:
        return match.group(1).strip(), match.group(2)
    return name.strip(), None


def fuzzy_match_roman_only(source_name: str, reference_names: List[str], max_distance: int = 3) -> Tuple[Optional[str], Optional[int]]:
    """Match source name to reference names using Roman numeral normalization."""
    if not source_name:
        return None, None

    no_num_max_dist = min(max_distance, 2)
    norm_source = normalize_name(source_name)
    arabic_source = roman_to_arabic(norm_source)
    source_base, source_num = _split_name_and_number(arabic_source)

    best_name = None
    best_dist = max_distance + 1

    for ref_name in reference_names:
        norm_ref = normalize_name(ref_name)
        if not norm_ref:
            continue

        arabic_ref = roman_to_arabic(norm_ref)
        ref_base, ref_num = _split_name_and_number(arabic_ref)

        if source_num is not None and ref_num is not None:
            if source_num != ref_num:
                continue
            dist = levenshtein_distance(source_base, ref_base)
        elif source_num is not None and ref_num is None:
            dist = levenshtein_distance(arabic_source, arabic_ref)
        elif source_num is None and ref_num is not None:
            dist = levenshtein_distance(arabic_source, arabic_ref)
        else:
            dist = levenshtein_distance(arabic_source, arabic_ref)
            if dist > no_num_max_dist:
                continue

        if dist < best_dist:
            best_dist = dist
            best_name = ref_name

    if best_dist <= max_distance:
        return best_name, best_dist
    return None, None


def fuzzy_match_all(source_name: str, reference_names: List[str], max_distance: int = 3) -> List[Tuple[str, int, str]]:
    """Find all candidate matches within Levenshtein max_distance."""
    if not source_name:
        return []

    norm_source = normalize_name(source_name)
    roman_source = roman_to_arabic(norm_source)
    arabic_source = arabic_to_roman(norm_source)

    src_roman_base, src_roman_num = _split_name_and_number(roman_source)
    src_arabic_base, src_arabic_num = _split_name_and_number(arabic_source)

    candidates = []
    seen = set()

    for ref_name in reference_names:
        norm_ref = normalize_name(ref_name)
        if not norm_ref:
            continue

        dist = levenshtein_distance(norm_source, norm_ref)
        method = 'DIRECT'

        roman_ref = roman_to_arabic(norm_ref)
        ref_roman_base, ref_roman_num = _split_name_and_number(roman_ref)

        if src_roman_num is not None and ref_roman_num is not None:
            if src_roman_num == ref_roman_num:
                dist_roman = levenshtein_distance(src_roman_base, ref_roman_base)
            else:
                dist_roman = max_distance + 1
        else:
            dist_roman = levenshtein_distance(roman_source, roman_ref)

        if dist_roman < dist:
            dist = dist_roman
            method = 'ROMAN_NUMERAL'

        arabic_ref = arabic_to_roman(norm_ref)
        ref_arabic_base, ref_arabic_num = _split_name_and_number(arabic_ref)

        if src_arabic_num is not None and ref_arabic_num is not None:
            if src_arabic_num == ref_arabic_num:
                dist_arabic = levenshtein_distance(src_arabic_base, ref_arabic_base)
            else:
                dist_arabic = max_distance + 1
        else:
            dist_arabic = levenshtein_distance(arabic_source, arabic_ref)

        if dist_arabic < dist:
            dist = dist_arabic
            method = 'ROMAN_NUMERAL'

        if dist <= max_distance and ref_name not in seen:
            candidates.append((ref_name, dist, method))
            seen.add(ref_name)

    candidates.sort(key=lambda x: x[1])
    return candidates


def title_case_smart(name: Any) -> str:
    """Smart title-case preserving Roman numerals."""
    if name is None or name == NULL or str(name).strip() == '':
        return ''

    text = str(name).strip()
    all_caps_pattern = r'^([(-]?(?:[A-Z][^ ]*|[IVXLCDM]+))( [(-]?(?:[A-Z][^ ]*|[IVXLCDM]+))*$'
    strict_caps_pattern = r'^([(-]?[A-Z]+)([ -()][(-]?[A-Z]+)*$'

    if re.match(all_caps_pattern, text):
        if re.match(strict_caps_pattern, text):
            return text.lower().title()
        else:
            return text
    else:
        return text.lower().title()


# ─── Processing Algorithm Implementation ───────────────────────────────────

class JoinBarangayAttributes(QgsProcessingAlgorithm):
    """
    Join Barangay Attributes Processing Algorithm.
    Enhanced fuzzy matching against embedded PSGC Q4.xlsx reference table.
    """

    INPUT = "citymun"
    FIELD = "field"
    MAX_DISTANCE = "max_distance"
    OUTPUT = "Bgy_name"
    OUTPUT_PSGC = "Filtered_PSGC"

    def name(self) -> str:
        return "join_barangay_attributes"

    def displayName(self) -> str:
        return "Join Barangay Attributes"

    def group(self) -> str:
        return "1Map"

    def groupId(self) -> str:
        return "1map"

    def icon(self):
        icon_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'icons', 'upload.svg')
        if os.path.exists(icon_path):
            return QIcon(icon_path)
        return QIcon(":/images/themes/default/mActionFilter.svg")

    def shortHelpString(self) -> str:
        return (
            "Enhanced Join Barangay Attributes with multi-stage fuzzy matching.\n\n"
            "This algorithm matches and standardizes barangay names from an input city/municipality "
            "vector layer against the embedded official PSGC reference dataset (references/PSGC Q4.xlsx).\n\n"
            "Matching Engine Workflow:\n"
            "  1. Exact Match — Direct 1:1 case-insensitive matching against the reference dataset.\n"
            "  2. Roman Numeral Normalization — Converts Roman numerals to Arabic numbers "
            "(e.g., 'Poblacion III' ↔ 'Poblacion 3') to pair equivalent names automatically.\n"
            "  3. Levenshtein Fuzzy Matching — Computes character edit distance (threshold = 3) "
            "for remaining unmatched barangays.\n\n"
            "Output Layer Columns:\n"
            "  • <field_name> — Original barangay attribute from source layer (converted to smart Title Case)\n"
            "  • barangay name (Final Name) — Resolved official PSGC barangay name\n"
            "  • error_detail — Diagnostic audit description (empty on clean matches; reports "
            "Roman/Arabic transformations, ambiguous candidate collisions, or unmatched records)\n\n"
            "Outputs Generated:\n"
            "  • Matched Barangays — Resulting vector layer with original geometries, standard names, and audit trail\n"
            "  • Filtered PSGC Table — Non-spatial attribute table of all reference PSGC records for the target LGU"
        )

    def initAlgorithm(self, config: Optional[Dict[str, Any]] = None):
        # 1. City/Municipality vector layer
        self.addParameter(
            QgsProcessingParameterFeatureSource(
                self.INPUT,
                "city/mun",
                [QgsProcessing.SourceType.TypeVectorPolygon, QgsProcessing.SourceType.TypeVectorAnyGeometry],
            )
        )

        # 2. Barangay attribute field
        self.addParameter(
            QgsProcessingParameterField(
                self.FIELD,
                "field",
                parentLayerParameterName=self.INPUT,
                type=QgsProcessingParameterField.Any,
            )
        )

        # 3. Max Levenshtein Distance (hidden, fixed at 3)
        max_dist_param = QgsProcessingParameterNumber(
            self.MAX_DISTANCE,
            "Max Levenshtein Distance",
            type=QgsProcessingParameterNumber.Integer,
            defaultValue=3,
            minValue=1,
            maxValue=10,
        )
        max_dist_param.setFlags(max_dist_param.flags() | QgsProcessingParameterDefinition.FlagHidden)
        self.addParameter(max_dist_param)

        # 4. Output Feature Sink: Matched Barangays
        self.addParameter(
            QgsProcessingParameterFeatureSink(self.OUTPUT, "Matched Barangays")
        )

        # 5. Output Feature Sink: Filtered PSGC Table
        self.addParameter(
            QgsProcessingParameterFeatureSink(
                self.OUTPUT_PSGC,
                "Filtered PSGC Table",
                type=QgsProcessing.TypeVector,
            )
        )

    def processAlgorithm(
        self,
        parameters: Dict[str, Any],
        context: QgsProcessingContext,
        feedback: QgsProcessingFeedback,
    ) -> Dict[str, Any]:

        source = self.parameterAsSource(parameters, self.INPUT, context)
        if source is None:
            raise QgsProcessingException(self.invalidSourceError(parameters, self.INPUT))

        field_name = self.parameterAsString(parameters, self.FIELD, context)
        max_distance = self.parameterAsInt(parameters, self.MAX_DISTANCE, context)

        field_names = [f.name() for f in source.fields()]
        if field_name not in field_names:
            raise QgsProcessingException(f"Field '{field_name}' not found in input layer.")

        # Determine layer name and 5-digit code filter
        input_layer = QgsProcessingUtils.mapLayerFromString(str(parameters.get(self.INPUT)), context)
        citymun_name = input_layer.name() if input_layer and input_layer.isValid() else ""
        if not citymun_name and hasattr(source, "sourceName"):
            src_n = source.sourceName()
            citymun_name = str(src_n) if src_n is not None else ""
        if not citymun_name:
            citymun_name = "LGU_Layer"

        citymun_name = str(citymun_name)
        code_filter = str(citymun_name[:5]) if len(citymun_name) >= 5 else citymun_name

        # ── Resolve Built-in PSGC Excel File ────────────────────────
        psgc_file_path = os.path.abspath(
            os.path.join(os.path.dirname(os.path.dirname(__file__)), "references", "PSGC Q4.xlsx")
        )
        if not os.path.exists(psgc_file_path):
            rel_path = os.path.abspath(os.path.join(os.getcwd(), "references", "PSGC Q4.xlsx"))
            if os.path.exists(rel_path):
                psgc_file_path = rel_path
            else:
                raise QgsProcessingException(f"Built-in PSGC reference file not found at: '{psgc_file_path}'")

        feedback.pushInfo(f"Reading embedded PSGC reference file: {psgc_file_path}...")

        # ── Read PSGC Records (Matching update_metadata_by_geocode pattern) ──
        all_records = []
        target_cols = [
            "map_uuid", "geocode", "region", "province", "city_mun", "barangay",
            "province_code", "city_mun_code", "barangay_code", "region_code",
            "office_region", "office_pso", "hhcount", "bldgcount"
        ]

        def _clean_segment(v, length=0):
            if v is None or v == NULL:
                return ""
            s = str(v).strip()
            if s.endswith(".0"):
                s = s[:-2]
            s = "".join(c for c in s if c.isdigit())
            if s and length > 0:
                s = s.zfill(length)
            return s

        def _parse_int(v):
            if v is None or v == NULL:
                return None
            s = str(v).strip()
            if s == "" or s.lower() in ("nan", "none", "null"):
                return None
            try:
                return int(float(s))
            except (ValueError, TypeError):
                return None

        # Method 1: QgsVectorLayer via GDAL/OGR
        psgc_layer = QgsVectorLayer(f"{psgc_file_path}|layername=PSGC", "psgc_ref", "ogr")
        if not psgc_layer or not psgc_layer.isValid():
            psgc_layer = QgsVectorLayer(psgc_file_path, "psgc_ref", "ogr")

        if psgc_layer and psgc_layer.isValid():
            fields = psgc_layer.fields()
            col_indices = {}
            for col_name in target_cols:
                col_norm = col_name.replace("_", "").lower()
                for field in fields:
                    if field.name().lower().replace("_", "").replace("/", "") == col_norm:
                        col_indices[col_name] = field.name()
                        break

            if "geocode" in col_indices or "province_code" in col_indices:
                for feat in psgc_layer.getFeatures():
                    if feedback.isCanceled():
                        break

                    prov_c = _clean_segment(feat.attribute(col_indices["province_code"])) if "province_code" in col_indices else ""
                    city_c = _clean_segment(feat.attribute(col_indices["city_mun_code"]), 2) if "city_mun_code" in col_indices else ""
                    bgy_c = _clean_segment(feat.attribute(col_indices["barangay_code"]), 3) if "barangay_code" in col_indices else ""

                    lgu_c = f"{prov_c}{city_c}" if (prov_c and city_c) else ""

                    def _get_val(k):
                        if k in col_indices:
                            v = feat.attribute(col_indices[k])
                            return str(v).strip() if v is not None and v != NULL else ""
                        return ""

                    raw_geo = _get_val("geocode")
                    geo_clean = "".join(c for c in raw_geo.split(".")[0] if c.isdigit())

                    rec = {
                        "map_uuid": _get_val("map_uuid"),
                        "geocode": geo_clean,
                        "lgu_code": lgu_c,
                        "province_code": prov_c,
                        "city_mun_code": city_c,
                        "barangay_code": bgy_c,
                        "region_code": _get_val("region_code"),
                        "region": _get_val("region"),
                        "province": _get_val("province"),
                        "city_mun": _get_val("city_mun"),
                        "barangay": _get_val("barangay"),
                        "office_region": _get_val("office_region"),
                        "office_pso": _get_val("office_pso"),
                        "hhcount": _parse_int(feat.attribute(col_indices["hhcount"])) if "hhcount" in col_indices else None,
                        "bldgcount": _parse_int(feat.attribute(col_indices["bldgcount"])) if "bldgcount" in col_indices else None,
                    }
                    all_records.append(rec)

        # Method 2: openpyxl fallback
        if not all_records:
            try:
                wb = openpyxl.load_workbook(psgc_file_path, read_only=True, data_only=True)
                sheet_name = "PSGC" if "PSGC" in wb.sheetnames else wb.sheetnames[0]
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)
                header_row = next(rows_iter)
                header_clean = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]

                col_indices = {}
                for col_name in target_cols:
                    if col_name in header_clean:
                        col_indices[col_name] = header_clean.index(col_name)

                if "geocode" in col_indices or "province_code" in col_indices:
                    for row in rows_iter:
                        if feedback.isCanceled():
                            break

                        prov_c = _clean_segment(row[col_indices["province_code"]]) if "province_code" in col_indices else ""
                        city_c = _clean_segment(row[col_indices["city_mun_code"]], 2) if "city_mun_code" in col_indices else ""
                        bgy_c = _clean_segment(row[col_indices["barangay_code"]], 3) if "barangay_code" in col_indices else ""

                        lgu_c = f"{prov_c}{city_c}" if (prov_c and city_c) else ""

                        def _row_val(k):
                            if k in col_indices and row[col_indices[k]] is not None:
                                return str(row[col_indices[k]]).strip()
                            return ""

                        raw_geo = _row_val("geocode")
                        geo_clean = "".join(c for c in raw_geo.split(".")[0] if c.isdigit())

                        rec = {
                            "map_uuid": _row_val("map_uuid"),
                            "geocode": geo_clean,
                            "lgu_code": lgu_c,
                            "province_code": prov_c,
                            "city_mun_code": city_c,
                            "barangay_code": bgy_c,
                            "region_code": _row_val("region_code"),
                            "region": _row_val("region"),
                            "province": _row_val("province"),
                            "city_mun": _row_val("city_mun"),
                            "barangay": _row_val("barangay"),
                            "office_region": _row_val("office_region"),
                            "office_pso": _row_val("office_pso"),
                            "hhcount": _parse_int(row[col_indices["hhcount"]]) if "hhcount" in col_indices else None,
                            "bldgcount": _parse_int(row[col_indices["bldgcount"]]) if "bldgcount" in col_indices else None,
                        }
                        all_records.append(rec)
                wb.close()
            except Exception as e:
                raise QgsProcessingException(f"Failed to read PSGC file: {e}")

        # ── Filter Records for Target City/Municipality ──────────────
        filtered_records = []
        if code_filter:
            filtered_records = [r for r in all_records if str(r.get("lgu_code", "")) == code_filter or str(r.get("geocode", "")).startswith(code_filter)]

        if not filtered_records and citymun_name:
            clean_layer = citymun_name.lower().replace("_", " ")
            for r in all_records:
                city = r.get("city_mun", "")
                if city and (city.lower() in clean_layer or clean_layer in city.lower()):
                    filtered_records.append(r)

        actual_city_name = filtered_records[0].get("city_mun", "") if filtered_records else ""
        if not actual_city_name:
            raw_name = citymun_name[5:].strip() if len(citymun_name) > 5 else citymun_name
            actual_city_name = raw_name if raw_name else "Unknown"

        feedback.pushInfo(f"Filtered {len(filtered_records)} PSGC barangay records for '{actual_city_name}'.")

        # ── Populate Filtered PSGC Table Sink ───────────────────────
        psgc_fields = QgsFields()
        psgc_cols = [
            "map_uuid", "geocode", "lgu_code", "region_code", "province_code",
            "city_mun_code", "barangay_code", "region", "province", "city_mun",
            "barangay", "office_region", "office_pso", "hhcount", "bldgcount"
        ]
        for col in psgc_cols:
            if col in ("hhcount", "bldgcount"):
                psgc_fields.append(QgsField(col, QVariant.Int))
            else:
                psgc_fields.append(QgsField(col, QVariant.String, len=100))

        (psgc_sink, psgc_dest_id) = self.parameterAsSink(
            parameters,
            self.OUTPUT_PSGC,
            context,
            psgc_fields,
            getattr(QgsWkbTypes, 'NoGeometry', getattr(QgsWkbTypes, 'NullGeometry', 0)),
            source.sourceCrs(),
        )

        if psgc_sink is not None:
            for r in filtered_records:
                if feedback.isCanceled():
                    break
                f_psgc = QgsFeature(psgc_fields)
                for col in psgc_cols:
                    val = r.get(col)
                    if col in ("hhcount", "bldgcount"):
                        f_psgc.setAttribute(col, val if (val is not None and val != "" and val != NULL) else NULL)
                    else:
                        f_psgc.setAttribute(col, str(val) if (val is not None and val != NULL) else "")
                psgc_sink.addFeature(f_psgc, QgsFeatureSink.FastInsert)

        # ── Build Reference Maps ────────────────────────────────────
        reference_names = []
        exact_match_map = {}
        for r in filtered_records:
            bgy = r.get("barangay", "").strip()
            if bgy and bgy != 'NULL':
                reference_names.append(bgy)
                exact_match_map[normalize_name(bgy)] = bgy

        # ── Create Output Feature Sink ──────────────────────────────
        source_field_def = None
        source_field_idx = source.fields().indexOf(field_name)
        if source_field_idx != -1:
            source_field_def = source.fields().at(source_field_idx)

        output_fields = QgsFields()
        if source_field_def:
            output_fields.append(QgsField(source_field_def.name(), source_field_def.type(),
                                          source_field_def.typeName(), source_field_def.length(),
                                          source_field_def.precision()))
        else:
            output_fields.append(QgsField(field_name, QVariant.String, len=254))

        output_fields.append(QgsField('barangay name (Final Name)', QVariant.String, len=254))
        output_fields.append(QgsField('error_detail', QVariant.String, len=500))

        (sink, dest_id) = self.parameterAsSink(
            parameters,
            self.OUTPUT,
            context,
            output_fields,
            source.wkbType(),
            source.sourceCrs(),
        )
        if sink is None:
            raise QgsProcessingException(self.invalidSinkError(parameters, self.OUTPUT))

        self.dest_id = dest_id
        self.psgc_dest_id = psgc_dest_id

        if code_filter:
            self.custom_name = f"{code_filter}_{actual_city_name} (Matched)"
            self.psgc_custom_name = f"{code_filter}_{actual_city_name} (Filtered PSGC)"
        else:
            self.custom_name = f"{actual_city_name} (Matched)"
            self.psgc_custom_name = f"{actual_city_name} (Filtered PSGC)"

        # ── Process Features & Match with Match Cache Memoization ───
        stats = {
            'total': 0, 'exact': 0, 'fuzzy': 0,
            'multiple': 0, 'roman': 0, 'no_match': 0
        }

        # Cache key: source_name string -> (final_name, error_detail, stat_type)
        match_cache: Dict[str, Tuple[str, str, str]] = {}

        total_feats = source.featureCount()
        step = 100.0 / total_feats if total_feats else 0
        current = 0

        for feat in source.getFeatures():
            if feedback.isCanceled():
                break

            stats['total'] += 1
            current += 1

            raw_val = feat.attribute(field_name)
            source_name = str(raw_val).strip() if raw_val is not None and raw_val != NULL else ""
            title_cased = title_case_smart(source_name)

            out_feat = QgsFeature(output_fields)
            out_feat.setGeometry(feat.geometry())
            out_feat.setAttribute(field_name, title_cased if title_cased else source_name)

            # Check match cache first to bypass expensive Levenshtein recalculations
            if source_name in match_cache:
                final_name, error_detail, stat_type = match_cache[source_name]
                out_feat.setAttribute('barangay name (Final Name)', final_name)
                out_feat.setAttribute('error_detail', error_detail)
                stats[stat_type] += 1
            else:
                # 1. Exact match check
                norm_source = normalize_name(source_name)
                exact_match = exact_match_map.get(norm_source)

                # 2. Roman numeral check
                roman_match, roman_dist = fuzzy_match_roman_only(
                    source_name, reference_names, max_distance)

                has_exact = bool(exact_match)
                final_name = exact_match if has_exact else (roman_match if roman_match else "")
                error_detail = ""
                stat_type = "exact"

                if has_exact:
                    stat_type = "exact"
                else:
                    candidates = fuzzy_match_all(
                        source_name, reference_names, max_distance)

                    if not candidates:
                        error_detail = f'No match found for "{source_name}"'
                        stat_type = "no_match"
                    elif len(candidates) == 1:
                        best_name, best_dist, method = candidates[0]
                        if method == 'ROMAN_NUMERAL':
                            error_detail = (
                                f'Matched via Roman/Arabic normalization: '
                                f'"{source_name}" → "{best_name}" (dist={best_dist})'
                            )
                            stat_type = "roman"
                        else:
                            stat_type = "fuzzy"
                    else:
                        best_name, best_dist, method = candidates[0]
                        same_dist = [c for c in candidates if c[1] == best_dist]

                        if len(same_dist) > 1:
                            error_detail = (
                                f'{len(same_dist)} candidates with same distance '
                                f'{best_dist} for "{source_name}" — review needed'
                            )
                            stat_type = "multiple"
                        elif method == 'ROMAN_NUMERAL':
                            error_detail = (
                                f'Matched via Roman/Arabic normalization: '
                                f'"{source_name}" → "{best_name}" (dist={best_dist})'
                            )
                            stat_type = "roman"
                        else:
                            stat_type = "fuzzy"

                match_cache[source_name] = (final_name, error_detail, stat_type)
                out_feat.setAttribute('barangay name (Final Name)', final_name)
                out_feat.setAttribute('error_detail', error_detail)
                stats[stat_type] += 1

            sink.addFeature(out_feat, QgsFeatureSink.FastInsert)
            if step > 0:
                feedback.setProgress(int(current * step))

        feedback.pushInfo('')
        feedback.pushInfo('━' * 45)
        feedback.pushInfo('  FUZZY MATCH SUMMARY')
        feedback.pushInfo('━' * 45)
        feedback.pushInfo(f"  ✅ Exact matches:        {stats['exact']}")
        feedback.pushInfo(f"  🔄 Fuzzy matches:        {stats['fuzzy']}")
        feedback.pushInfo(f"  🔢 Roman numeral fixes:  {stats['roman']}")
        feedback.pushInfo(f"  ⚠️ Multiple matches:     {stats['multiple']}")
        feedback.pushInfo(f"  ❌ No match found:        {stats['no_match']}")
        feedback.pushInfo('━' * 45)
        feedback.pushInfo(f"  Total features:          {stats['total']}")
        feedback.pushInfo('━' * 45)

        results = {self.OUTPUT: dest_id}
        if psgc_dest_id:
            results[self.OUTPUT_PSGC] = psgc_dest_id
        return results

    def postProcessAlgorithm(
        self,
        context: QgsProcessingContext,
        feedback: QgsProcessingFeedback,
    ) -> Dict[str, Any]:
        matched_dest_id = getattr(self, "dest_id", None)
        matched_name = getattr(self, "custom_name", "Matched_Barangays")
        psgc_dest_id = getattr(self, "psgc_dest_id", None)
        psgc_name = getattr(self, "psgc_custom_name", "Filtered_PSGC")

        # 1. Clean up duplicate entries in layersToLoadOnCompletion
        layers_to_load = context.layersToLoadOnCompletion()
        if layers_to_load:
            new_layers_to_load = {}
            has_specific_psgc = bool(psgc_dest_id and psgc_dest_id in layers_to_load)
            has_specific_matched = bool(matched_dest_id and matched_dest_id in layers_to_load)

            for k, details in layers_to_load.items():
                is_psgc = (
                    details.outputName == self.OUTPUT_PSGC
                    or details.name in ("Filtered PSGC Table", "Filtered_PSGC", psgc_name)
                    or (psgc_dest_id and k == psgc_dest_id)
                )
                is_matched = (
                    details.outputName == self.OUTPUT
                    or details.name in ("Matched Barangays", "Bgy_name", matched_name)
                    or (matched_dest_id and k == matched_dest_id)
                )

                if is_psgc:
                    # Skip generic placeholder if a concrete sink dest_id is present
                    if has_specific_psgc and k != psgc_dest_id:
                        continue
                    details.name = psgc_name
                    new_layers_to_load[k] = details
                elif is_matched:
                    # Skip generic placeholder if a concrete sink dest_id is present
                    if has_specific_matched and k != matched_dest_id:
                        continue
                    details.name = matched_name
                    new_layers_to_load[k] = details
                else:
                    new_layers_to_load[k] = details

            try:
                context.setLayersToLoadOnCompletion(new_layers_to_load)
            except Exception:
                pass

        # 2. Update specific dest_id details if present
        if matched_dest_id and context.willLoadLayerOnCompletion(matched_dest_id):
            details = context.layerToLoadOnCompletionDetails(matched_dest_id)
            details.name = matched_name

        if psgc_dest_id and context.willLoadLayerOnCompletion(psgc_dest_id):
            details = context.layerToLoadOnCompletionDetails(psgc_dest_id)
            details.name = psgc_name

        # 3. Synchronize names of any layers already loaded or mapped in context
        if psgc_dest_id:
            psgc_layer = QgsProcessingUtils.mapLayerFromString(psgc_dest_id, context)
            if psgc_layer and psgc_layer.isValid():
                psgc_layer.setName(psgc_name)

        if matched_dest_id:
            matched_layer = QgsProcessingUtils.mapLayerFromString(matched_dest_id, context)
            if matched_layer and matched_layer.isValid():
                matched_layer.setName(matched_name)

        # 4. Clean up any rogue unrenamed "Filtered PSGC Table" layers from the active project
        proj = context.project() if hasattr(context, "project") and context.project() else QgsProject.instance()
        if proj:
            for l in list(proj.mapLayers().values()):
                if l.name() == "Filtered PSGC Table":
                    proj.removeMapLayer(l.id())

        results = {}
        if matched_dest_id:
            results[self.OUTPUT] = matched_dest_id
        if psgc_dest_id:
            results[self.OUTPUT_PSGC] = psgc_dest_id
        return results

    def createInstance(self):
        return JoinBarangayAttributes()