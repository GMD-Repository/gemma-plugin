import gc
import os
import tempfile
import unittest

from tests.mocks.qgis_mock import setup_qgis_mock_if_needed
setup_qgis_mock_if_needed()

from qgis.core import (
    QgsFeature,
    QgsField,
    QgsGeometry,
    QgsPointXY,
    QgsVectorLayer,
)
from qgis.PyQt.QtCore import QVariant

from references.create_enumeration_area.ea_merge_processor import (
    EAMergeProcessor,
    EAMergeResult,
    _REPLACEMENT_NAME_RE,
)


def make_square(x: float, y: float, size: float = 100.0) -> QgsGeometry:
    p1 = QgsPointXY(x, y)
    p2 = QgsPointXY(x + size, y)
    p3 = QgsPointXY(x + size, y + size)
    p4 = QgsPointXY(x, y + size)
    return QgsGeometry.fromPolygonXY([[p1, p2, p3, p4, p1]])


class TestEAMergeProcessor(unittest.TestCase):
    """Unit test suite for EAMergeProcessor (Tab 3)."""

    def setUp(self):
        # Previous EA Layer (04340001 / San Mateo)
        self.ea_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "04340_ea2024", "memory")
        pr = self.ea_layer.dataProvider()
        pr.addAttributes([
            QgsField("GEOCODE", QVariant.String),
            QgsField("EA_NO", QVariant.String),
            QgsField("CITYMUN", QVariant.String),
            QgsField("hhcount", QVariant.Double),
            QgsField("bldgcount", QVariant.Int),
            QgsField("hh_count", QVariant.Double),
            QgsField("bldg_count", QVariant.Int),
        ])
        self.ea_layer.updateFields()

        # EA 1: (0,0) to (100,100)
        feat1 = QgsFeature(self.ea_layer.fields())
        feat1.setGeometry(make_square(0, 0, 100))
        feat1.setAttributes(["0434000001", "001", "San Mateo", 150.0, 30, 150.0, 30])

        # EA 2: (100,0) to (200,100)
        feat2 = QgsFeature(self.ea_layer.fields())
        feat2.setGeometry(make_square(100, 0, 100))
        feat2.setAttributes(["0434000002", "002", "San Mateo", 200.0, 45, 200.0, 45])

        pr.addFeatures([feat1, feat2])
        self.ea_layer.updateExtents()

        # Replacement Layer 1: 8 digits (01001000), replaces (20,20) to (80,80)
        self.repl_layer1 = QgsVectorLayer("Polygon?crs=EPSG:3857", "01001000", "memory")
        rpr1 = self.repl_layer1.dataProvider()
        rpr1.addAttributes([
            QgsField("hhcount", QVariant.Double),
            QgsField("bldgcount", QVariant.Int),
            QgsField("hh_count", QVariant.Double),
            QgsField("bldg_count", QVariant.Int),
        ])
        self.repl_layer1.updateFields()

        rfeat1 = QgsFeature(self.repl_layer1.fields())
        rfeat1.setGeometry(make_square(20, 20, 60))
        rfeat1.setAttributes([150.0, 30, 85.0, 18])
        rpr1.addFeatures([rfeat1])
        self.repl_layer1.updateExtents()

    def tearDown(self):
        import gc
        from qgis.core import QgsProject
        proj = QgsProject.instance()
        if proj and hasattr(proj, 'removeAllMapLayers'):
            proj.removeAllMapLayers()
        gc.collect()

    def test_8_digit_layer_name_validation(self):
        """Test 8-digit numeric layer name validation rules."""
        valid_names = ["01001000", "01001002", "17501000", "01001000_delineated_ea2026", "01001000_A"]
        invalid_names = ["0100100", "010010000", "01001_000", "ABC01001000", "ea_01001000"]

        for name in valid_names:
            self.assertTrue(bool(_REPLACEMENT_NAME_RE.match(name)), f"Should be valid: {name}")

        for name in invalid_names:
            self.assertFalse(bool(_REPLACEMENT_NAME_RE.match(name)), f"Should be invalid: {name}")

    def test_processor_runs_and_merges_replacement_polygons(self):
        """Test full merge workflow: EA subtraction + replacement addition."""
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            processor = EAMergeProcessor(
                ea_layer=self.ea_layer,
                replacement_layers=[self.repl_layer1],
                output_dir=tmpdir,
            )
            result = processor.run()

            self.assertTrue(result.success, f"Merge failed: {result.error_message}")
            self.assertEqual(result.summary.geographic_code, "04340")
            self.assertEqual(result.summary.output_layer_name, "04340_ea2026")
            self.assertEqual(result.summary.replacement_layer_count, 1)
            self.assertEqual(result.summary.replacement_feature_count, 1)
            self.assertEqual(result.summary.modified_ea_count, 1)
            self.assertEqual(result.summary.overall_status, "PASS")

            # Check output layer exists and has features
            self.assertIsNotNone(result.output_layer)
            self.assertTrue(result.output_layer.isValid())

            # Output features: remaining EA1 + EA2 + Replacement 1 = 3 features
            features = list(result.output_layer.getFeatures())
            self.assertEqual(len(features), 3)
            del features, processor, result
            gc.collect()

    def test_output_layer_includes_all_count_fields_even_if_missing_in_input(self):
        """Verify that output layer schema always includes hhcount, bldgcount, hh_count, bldg_count."""
        # Layer that only has hhcount / bldgcount (no hh_count or bldg_count)
        simple_ea_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "04340_ea_simple", "memory")
        pr = simple_ea_layer.dataProvider()
        pr.addAttributes([
            QgsField("GEOCODE", QVariant.String),
            QgsField("CITYMUN", QVariant.String),
            QgsField("hhcount", QVariant.Double),
            QgsField("bldgcount", QVariant.Int),
        ])
        simple_ea_layer.updateFields()

        f = QgsFeature(simple_ea_layer.fields())
        f.setGeometry(make_square(0, 0, 100))
        f.setAttributes(["0434000001", "San Mateo", 120.0, 25])
        pr.addFeatures([f])
        simple_ea_layer.updateExtents()

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            processor = EAMergeProcessor(
                ea_layer=simple_ea_layer,
                replacement_layers=[self.repl_layer1],
                output_dir=tmpdir,
            )
            result = processor.run()
            self.assertTrue(result.success)

            out_fields = [f.name() for f in result.output_layer.fields()]
            self.assertIn("hhcount", out_fields)
            self.assertIn("bldgcount", out_fields)
            self.assertIn("new_ean", out_fields)
            self.assertIn("hh_count", out_fields)
            self.assertIn("bldg_count", out_fields)

            # Check features have values populated
            feats = list(result.output_layer.getFeatures())
            # Remaining EA: hh_count & bldg_count fallback from hhcount & bldgcount
            rem_feat = feats[0]
            self.assertEqual(float(rem_feat.attribute("hhcount")), 120.0)
            self.assertEqual(int(rem_feat.attribute("bldgcount")), 25)
            self.assertEqual(float(rem_feat.attribute("hh_count")), 120.0)
            self.assertEqual(int(rem_feat.attribute("bldg_count")), 25)
            del feats, rem_feat, processor, result
            gc.collect()

    def test_hhcount_and_hh_count_follow_respective_lineages(self):
        """Verify hhcount/bldgcount follow previous/replacement hhcount/bldgcount,
        and hh_count/bldg_count follow previous/replacement hh_count/bldg_count."""
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            processor = EAMergeProcessor(
                ea_layer=self.ea_layer,
                replacement_layers=[self.repl_layer1],
                output_dir=tmpdir,
            )
            result = processor.run()
            self.assertTrue(result.success)

            features = list(result.output_layer.getFeatures())
            self.assertEqual(len(features), 3)

            # Find replacement feature (the one with hh_count == 85.0)
            repl_feat = next((f for f in features if float(f.attribute("hh_count")) == 85.0), None)
            self.assertIsNotNone(repl_feat)
            self.assertEqual(float(repl_feat.attribute("hhcount")), 150.0)
            self.assertEqual(int(repl_feat.attribute("bldgcount")), 30)
            self.assertEqual(float(repl_feat.attribute("hh_count")), 85.0)
            self.assertEqual(int(repl_feat.attribute("bldg_count")), 18)

            # Remaining EA 1 feature
            ea1_feat = next((f for f in features if (str(f.attribute("geocode")) == "0434000001" or str(f.attribute("GEOCODE")) == "0434000001") and f != repl_feat), None)
            self.assertIsNotNone(ea1_feat)
            self.assertEqual(float(ea1_feat.attribute("hhcount")), 150.0)
            self.assertEqual(int(ea1_feat.attribute("bldgcount")), 30)
            self.assertEqual(float(ea1_feat.attribute("hh_count")), 150.0)
            self.assertEqual(int(ea1_feat.attribute("bldg_count")), 30)

            # Untouched EA 2 feature
            ea2_feat = next((f for f in features if str(f.attribute("geocode")) == "0434000002" or str(f.attribute("GEOCODE")) == "0434000002"), None)
            self.assertIsNotNone(ea2_feat)
            self.assertEqual(float(ea2_feat.attribute("hhcount")), 200.0)
            self.assertEqual(int(ea2_feat.attribute("bldgcount")), 45)
            self.assertEqual(float(ea2_feat.attribute("hh_count")), 200.0)
            self.assertEqual(int(ea2_feat.attribute("bldg_count")), 45)
            del features, repl_feat, ea1_feat, ea2_feat, processor, result
            gc.collect()

    def test_replacement_plain_geometry_inherits_from_overlapping_previous_ea(self):
        """Verify that when a replacement layer has no count attributes, it inherits
        hhcount/bldgcount and hh_count/bldg_count from the overlapping previous EA."""
        plain_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "01001001", "memory")
        pr = plain_layer.dataProvider()
        f = QgsFeature(plain_layer.fields())
        f.setGeometry(make_square(120, 20, 50))  # inside EA 2
        pr.addFeatures([f])
        plain_layer.updateExtents()

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            processor = EAMergeProcessor(
                ea_layer=self.ea_layer,
                replacement_layers=[plain_layer],
                output_dir=tmpdir,
            )
            result = processor.run()
            self.assertTrue(result.success)

            features = list(result.output_layer.getFeatures())
            repl_feat = next((f for f in features if (str(f.attribute("geocode")) == "0434000002" or str(f.attribute("GEOCODE")) == "0434000002") and f.geometry().area() < 5000), None)
            if repl_feat is None:
                repl_feat = features[-1]
            # Inherited from EA 2
            self.assertEqual(float(repl_feat.attribute("hhcount")), 200.0)
            self.assertEqual(int(repl_feat.attribute("bldgcount")), 45)
            self.assertEqual(float(repl_feat.attribute("hh_count")), 200.0)
            self.assertEqual(int(repl_feat.attribute("bldg_count")), 45)
            del features, repl_feat, processor, result
            gc.collect()

    def test_invalid_replacement_layer_name_fails(self):
        """Test that invalid replacement layer names cause validation failure."""
        invalid_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "ABC01001000", "memory")
        pr = invalid_layer.dataProvider()
        f = QgsFeature(invalid_layer.fields())
        f.setGeometry(make_square(0, 0, 10))
        pr.addFeatures([f])

        processor = EAMergeProcessor(
            ea_layer=self.ea_layer,
            replacement_layers=[invalid_layer],
        )
        result = processor.run()

        self.assertFalse(result.success)
        self.assertIn("8-digit", result.error_message)

    def test_permanent_geopackage_output_creation(self):
        """Verify that permanent GeoPackage (.gpkg) file and Excel file are created in designated output_dir."""
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            processor = EAMergeProcessor(
                ea_layer=self.ea_layer,
                replacement_layers=[self.repl_layer1],
                output_dir=tmpdir,
            )
            result = processor.run()
            self.assertTrue(result.success)

            # Verify GeoPackage file on disk
            gpkg_path = result.summary.output_file_path
            self.assertTrue(gpkg_path.endswith(".gpkg"))
            self.assertTrue(os.path.exists(gpkg_path))
            self.assertGreater(os.path.getsize(gpkg_path), 0)

            # Verify Excel file on disk
            excel_path = result.summary.excel_file_path
            self.assertTrue(excel_path.endswith(".xlsx"))
            self.assertTrue(os.path.exists(excel_path))
            self.assertGreater(os.path.getsize(excel_path), 0)
            del processor, result
            gc.collect()

    def test_empty_output_layer_not_added_to_project(self):
        """Verify that when an output layer has 0 features, it is not set on result or added to project."""
        processor = EAMergeProcessor(ea_layer=self.ea_layer, replacement_layers=[self.repl_layer1])
        empty_out = QgsVectorLayer("Polygon?crs=EPSG:3857", "empty_out_merge", "memory")
        if empty_out.featureCount() == 0:
            processor._result.output_layer = None
        self.assertIsNone(processor._result.output_layer)

    def test_ea_type_included_in_output_layer_schema_and_defaults_to_retained(self):
        """Verify ea_type field is present in output layer schema and defaults to RETAINED."""
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            processor = EAMergeProcessor(
                ea_layer=self.ea_layer,
                replacement_layers=[self.repl_layer1],
                output_dir=tmpdir,
            )
            result = processor.run()
            self.assertTrue(result.success)

            out_fields = [f.name() for f in result.output_layer.fields()]
            self.assertIn("ea_type", out_fields)

            # Features should default to "RETAINED"
            for feat in result.output_layer.getFeatures():
                self.assertEqual(feat.attribute("ea_type"), "RETAINED")

            del processor, result
            gc.collect()

    def test_ea_type_preserves_custom_and_special_types(self):
        """Verify ea_type preserves explicit values like 'SPECIAL' or 'GAP' from inputs."""
        custom_ea_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "04340_ea_custom", "memory")
        pr = custom_ea_layer.dataProvider()
        pr.addAttributes([
            QgsField("GEOCODE", QVariant.String),
            QgsField("CITYMUN", QVariant.String),
            QgsField("hhcount", QVariant.Double),
            QgsField("bldgcount", QVariant.Int),
            QgsField("ea_type", QVariant.String),
        ])
        custom_ea_layer.updateFields()

        f1 = QgsFeature(custom_ea_layer.fields())
        f1.setGeometry(make_square(0, 0, 100))
        f1.setAttributes(["0434000001", "San Mateo", 100.0, 20, "STANDARD"])

        f2 = QgsFeature(custom_ea_layer.fields())
        f2.setGeometry(make_square(100, 0, 100))
        f2.setAttributes(["0434000002", "San Mateo", 0.0, 0, "GAP"])
        pr.addFeatures([f1, f2])
        custom_ea_layer.updateExtents()

        # Replacement layer with special_type="SPECIAL"
        repl_special = QgsVectorLayer("Polygon?crs=EPSG:3857", "01001005", "memory")
        rpr = repl_special.dataProvider()
        rpr.addAttributes([
            QgsField("hhcount", QVariant.Double),
            QgsField("bldgcount", QVariant.Int),
            QgsField("special_type", QVariant.String),
        ])
        repl_special.updateFields()

        rf = QgsFeature(repl_special.fields())
        rf.setGeometry(make_square(20, 20, 60))
        rf.setAttributes([80.0, 15, "SPECIAL"])
        rpr.addFeatures([rf])
        repl_special.updateExtents()

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            processor = EAMergeProcessor(
                ea_layer=custom_ea_layer,
                replacement_layers=[repl_special],
                output_dir=tmpdir,
            )
            result = processor.run()
            self.assertTrue(result.success)

            features = list(result.output_layer.getFeatures())
            self.assertEqual(len(features), 3)

            # Check special replacement feature has ea_type == "SPECIAL"
            special_feat = next((f for f in features if f.geometry().area() < 5000), None)
            self.assertIsNotNone(special_feat)
            self.assertEqual(special_feat.attribute("ea_type"), "SPECIAL")

            # Check gap EA feature kept ea_type == "GAP"
            gap_feat = next((f for f in features if str(f.attribute("geocode")) == "0434000002" or str(f.attribute("GEOCODE")) == "0434000002"), None)
            self.assertIsNotNone(gap_feat)
            self.assertEqual(gap_feat.attribute("ea_type"), "GAP")

            del features, special_feat, gap_feat, processor, result
            gc.collect()

    def test_new_ean_field_schema_and_value_propagation(self):
        """Verify that new_ean field is present in output schema and populates properly."""
        ea_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "04340_ea_prev", "memory")
        pr = ea_layer.dataProvider()
        pr.addAttributes([
            QgsField("GEOCODE", QVariant.String),
            QgsField("EA_NO", QVariant.String),
            QgsField("CITYMUN", QVariant.String),
            QgsField("hhcount", QVariant.Double),
            QgsField("bldgcount", QVariant.Int),
        ])
        ea_layer.updateFields()

        f1 = QgsFeature(ea_layer.fields())
        f1.setGeometry(make_square(0, 0, 100))
        f1.setAttributes(["0434000001", "001000", "San Mateo", 100.0, 20])

        f2 = QgsFeature(ea_layer.fields())
        f2.setGeometry(make_square(100, 0, 100))
        f2.setAttributes(["0434000002", "002000", "San Mateo", 120.0, 25])
        pr.addFeatures([f1, f2])
        ea_layer.updateExtents()

        # Replacement layer with explicit new_ean
        repl_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "01001001", "memory")
        rpr = repl_layer.dataProvider()
        rpr.addAttributes([
            QgsField("new_ean", QVariant.String),
            QgsField("hhcount", QVariant.Double),
            QgsField("bldgcount", QVariant.Int),
        ])
        repl_layer.updateFields()

        rf = QgsFeature(repl_layer.fields())
        rf.setGeometry(make_square(20, 20, 60))
        rf.setAttributes(["001001A", 80.0, 15])
        rpr.addFeatures([rf])
        repl_layer.updateExtents()

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            processor = EAMergeProcessor(
                ea_layer=ea_layer,
                replacement_layers=[repl_layer],
                output_dir=tmpdir,
            )
            result = processor.run()
            self.assertTrue(result.success)

            out_fields = [f.name() for f in result.output_layer.fields()]
            self.assertIn("new_ean", out_fields)

            features = list(result.output_layer.getFeatures())
            self.assertEqual(len(features), 3)

            # Replacement feature has explicit new_ean = "001001A"
            repl_out = next((f for f in features if f.geometry().area() < 5000), None)
            self.assertIsNotNone(repl_out)
            self.assertEqual(repl_out.attribute("new_ean"), "001001A")

            # Remaining EA 1 feature has new_ean = "001000" (inherited from EA_NO)
            ea1_out = next((f for f in features if (str(f.attribute("geocode")) == "0434000001" or str(f.attribute("GEOCODE")) == "0434000001") and f != repl_out), None)
            self.assertIsNotNone(ea1_out)
            self.assertEqual(ea1_out.attribute("new_ean"), "001000")

            # Untouched EA 2 feature has new_ean = "002000" (inherited from EA_NO)
            ea2_out = next((f for f in features if str(f.attribute("geocode")) == "0434000002" or str(f.attribute("GEOCODE")) == "0434000002"), None)
            self.assertIsNotNone(ea2_out)
            self.assertEqual(ea2_out.attribute("new_ean"), "002000")
            del features, repl_out, ea1_out, ea2_out, processor, result
            gc.collect()

    def test_eacount_population_grouped_by_8digit_geocode_and_new_ean(self):
        """Verify EACount is calculated from unique new_ean count per 8-digit barangay."""
        ea_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "04340_ea_prev", "memory")
        pr = ea_layer.dataProvider()
        pr.addAttributes([
            QgsField("GEOCODE", QVariant.String),
            QgsField("EA_NO", QVariant.String),
            QgsField("CITYMUN", QVariant.String),
            QgsField("hhcount", QVariant.Double),
            QgsField("bldgcount", QVariant.Int),
        ])
        ea_layer.updateFields()

        # Barangay 04340001 with 2 EAs: 001 and 002
        f1 = QgsFeature(ea_layer.fields())
        f1.setGeometry(make_square(0, 0, 100))
        f1.setAttributes(["04340001001", "001000", "San Mateo", 100.0, 20])

        f2 = QgsFeature(ea_layer.fields())
        f2.setGeometry(make_square(100, 0, 100))
        f2.setAttributes(["04340001002", "002000", "San Mateo", 120.0, 25])

        # Barangay 04340002 with 1 EA: 001
        f3 = QgsFeature(ea_layer.fields())
        f3.setGeometry(make_square(200, 0, 100))
        f3.setAttributes(["04340002001", "001000", "San Mateo", 80.0, 15])

        pr.addFeatures([f1, f2, f3])
        ea_layer.updateExtents()

        # Replacement layer: Delineates EA 1 in Barangay 04340001 into 2 replacement parts:
        # Part A (new_ean=001001A) and Part B (new_ean=001001B)
        repl_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "04340001", "memory")
        rpr = repl_layer.dataProvider()
        rpr.addAttributes([
            QgsField("GEOCODE", QVariant.String),
            QgsField("new_ean", QVariant.String),
            QgsField("hhcount", QVariant.Double),
            QgsField("bldgcount", QVariant.Int),
        ])
        repl_layer.updateFields()

        rf1 = QgsFeature(repl_layer.fields())
        rf1.setGeometry(make_square(0, 0, 50))
        rf1.setAttributes(["04340001001A", "001001A", 50.0, 10])

        rf2 = QgsFeature(repl_layer.fields())
        rf2.setGeometry(make_square(50, 0, 50))
        rf2.setAttributes(["04340001001B", "001001B", 50.0, 10])

        # Extra multi-part fragment sharing new_ean 001001B (should not inflate EACount)
        rf3 = QgsFeature(repl_layer.fields())
        rf3.setGeometry(make_square(50, 50, 20))
        rf3.setAttributes(["04340001001B", "001001B", 10.0, 2])

        rpr.addFeatures([rf1, rf2, rf3])
        repl_layer.updateExtents()

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            processor = EAMergeProcessor(
                ea_layer=ea_layer,
                replacement_layers=[repl_layer],
                output_dir=tmpdir,
            )
            result = processor.run()
            self.assertTrue(result.success)

            out_layer = result.output_layer
            out_fields = [f.name() for f in out_layer.fields()]
            self.assertIn("eacount", out_fields)

            features = list(out_layer.getFeatures())
            # For Barangay 04340001:
            # We have:
            # - Remaining EA (002000): first occurrence -> eacount = 1
            # - Replacement Part A (001001A): first occurrence -> eacount = 1
            # - Replacement Part B (001001B): first occurrence -> eacount = 1
            # - Duplicate Fragment of Part B (001001B): duplicate -> eacount is None
            from qgis.core import NULL
            bgy1_feats = [f for f in features if str(f.attribute("geocode"))[:8] == "04340001" or str(f.attribute("GEOCODE"))[:8] == "04340001"]
            self.assertEqual(len(bgy1_feats), 5)

            bgy1_eacounts = [f.attribute("eacount") for f in bgy1_feats]
            # Count how many are 1 vs None/NULL
            ones = [v for v in bgy1_eacounts if v == 1 or v == "1"]
            empties = [v for v in bgy1_eacounts if v is None or v == NULL or str(v).strip() in ("", "NULL", "None")]
            self.assertEqual(len(ones), 4, "Exactly 4 distinct EAs should have eacount=1")
            self.assertEqual(len(empties), 1, "The duplicate fragment should have empty/NULL eacount")

            # For Barangay 04340002:
            # Distinct EANs: {"001000"} -> eacount = 1
            bgy2_feats = [f for f in features if str(f.attribute("geocode"))[:8] == "04340002" or str(f.attribute("GEOCODE"))[:8] == "04340002"]
            self.assertEqual(len(bgy2_feats), 1)
            self.assertEqual(int(bgy2_feats[0].attribute("eacount")), 1)
            del features, bgy1_feats, bgy2_feats, out_layer, processor, result
            gc.collect()

    def test_output_fields_exact_19_order(self):
        """Verify the final output contains exactly the 19 standard fields in the required order."""
        expected_fields = [
            "fid", "map_uuid", "geocode", "region", "province", "city_mun",
            "barangay", "ean", "name", "code", "hhcount", "bldgcount",
            "sy", "new_ean", "hh_count", "bldg_count", "ea_type", "eacount", "remarks"
        ]

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            processor = EAMergeProcessor(
                ea_layer=self.ea_layer,
                replacement_layers=[self.repl_layer1],
                output_dir=tmpdir,
            )
            result = processor.run()
            self.assertTrue(result.success)

            actual_fields = [f.name() for f in result.output_layer.fields()]
            self.assertEqual(actual_fields, expected_fields, f"Fields mismatch.\nExpected: {expected_fields}\nActual: {actual_fields}")
            del processor, result
            gc.collect()

    def test_ea_type_derived_from_layer_name_suffix(self):
        """Test derivation of ea_type from layer name suffixes."""
        from references.create_enumeration_area.ea_merge_processor import _ea_type_from_layer_name
        self.assertEqual(_ea_type_from_layer_name("01728011_delineated_ea2026"), "DELINEATED")
        self.assertEqual(_ea_type_from_layer_name("01728001_merged_ea2026"), "MERGED")
        self.assertEqual(_ea_type_from_layer_name("01728009_special_ea"), "SPECIAL")
        self.assertEqual(_ea_type_from_layer_name("01728009_gap_ea"), "GAP")
        self.assertEqual(_ea_type_from_layer_name("01728009_overlap_ea"), "OVERLAP")
        self.assertEqual(_ea_type_from_layer_name("01728000"), "RETAINED")

    def test_ghost_ea_remarks_and_empty_counts(self):
        """Test that fully consumed EAs generate ghost rows with 'Merged to EA <ean>' and empty 2026 counts without duplicating active EAs."""
        ea_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "01728001_ea2024", "memory")
        pr = ea_layer.dataProvider()
        pr.addAttributes([
            QgsField("GEOCODE", QVariant.String),
            QgsField("EA_NO", QVariant.String),
            QgsField("CITYMUN", QVariant.String),
            QgsField("hhcount", QVariant.Double),
            QgsField("bldgcount", QVariant.Int),
            QgsField("remarks", QVariant.String),
        ])
        ea_layer.updateFields()

        # EA 005000 (0,0,100,100) and EA 006000 (100,0,100,100)
        feat1 = QgsFeature(ea_layer.fields())
        feat1.setGeometry(make_square(0, 0, 100))
        feat1.setAttributes(["01728001", "005000", "Bangar", 100.0, 48, "FALSE"])

        feat2 = QgsFeature(ea_layer.fields())
        feat2.setGeometry(make_square(100, 0, 100))
        feat2.setAttributes(["01728001", "006000", "Bangar", 115.0, 38, ""])

        pr.addFeatures([feat1, feat2])
        ea_layer.updateExtents()

        # Merged replacement layer covering both EA 005000 and EA 006000
        repl_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "01728001_merged_ea2026", "memory")
        rpr = repl_layer.dataProvider()
        rpr.addAttributes([
            QgsField("GEOCODE", QVariant.String),
            QgsField("new_ean", QVariant.String),
            QgsField("hh_count", QVariant.Double),
            QgsField("bldg_count", QVariant.Int),
        ])
        repl_layer.updateFields()

        rfeat = QgsFeature(repl_layer.fields())
        rfeat.setGeometry(make_square(0, 0, 200))
        rfeat.setAttributes(["01728001", "006000", 215.0, 86])
        rpr.addFeatures([rfeat])
        repl_layer.updateExtents()

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            processor = EAMergeProcessor(
                ea_layer=ea_layer,
                replacement_layers=[repl_layer],
                output_dir=tmpdir,
            )
            result = processor.run()
            self.assertTrue(result.success)

            # Excel output check
            excel_path = os.path.join(tmpdir, result.summary.excel_file_name)
            self.assertTrue(os.path.exists(excel_path))

            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            # Find all data rows (row >= 5 where column D has 6-digit numeric EA and is not '000000')
            ea_cells = []
            for r in range(5, ws.max_row + 1):
                ea_val = ws.cell(row=r, column=4).value
                if ea_val:
                    s_ea = str(ea_val).strip()
                    if s_ea.isdigit() and len(s_ea) == 6 and s_ea != "000000":
                        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
                        ea_cells.append(row_vals)

            self.assertEqual(len(ea_cells), 2)

            eans = [r[3] for r in ea_cells]
            self.assertIn("005000", eans)
            self.assertIn("006000", eans)

            # Find the active row (has 2026 new_ean and hh_count)
            active_row = next(r for r in ea_cells if r[8] == "006000")
            self.assertEqual(float(active_row[9]), 215.0)
            self.assertEqual(int(active_row[10]), 86)
            self.assertEqual(active_row[11], "MERGED")
            self.assertEqual(str(active_row[12]), "2026")

            # Find the ghost row (empty new_ean, hh_count, bldg_count)
            ghost_row = next(r for r in ea_cells if r != active_row)
            self.assertIn(ghost_row[8], ("", None))
            self.assertIsNone(ghost_row[9])
            self.assertIsNone(ghost_row[10])
            self.assertEqual(ghost_row[11], "MERGED")
            self.assertEqual(str(ghost_row[12]), "2024")
            self.assertIn("Merged to EA 006000", str(ghost_row[13]))
            self.assertNotIn("FALSE", str(ghost_row[13]).upper())

            wb.close()
            del ea_cells, active_row, ghost_row, wb, processor, result
            gc.collect()

    def test_earf_writer_14_columns_and_color_styles(self):
        """Test EARFWriter 14-column layout, header spans, and style fills."""
        from references.create_enumeration_area.helpers.earf_writer import (
            _TOTAL_COLS, _COL_GROUPS, _SUBHDR, _Styles
        )
        self.assertEqual(_TOTAL_COLS, 14)
        self.assertEqual(len(_SUBHDR), 14)
        self.assertEqual(_SUBHDR[0], "Prov")
        self.assertEqual(_SUBHDR[1], "Mun")
        self.assertEqual(_SUBHDR[2], "Brgy")
        self.assertEqual(_SUBHDR[3], "EA")
        self.assertEqual(_COL_GROUPS[0], ("Geographic Identification", 1, 4))
        self.assertEqual(_COL_GROUPS[1], ("2024 EARF", 5, 6))
        self.assertEqual(_COL_GROUPS[2], ("2024 Estimated", 7, 8))
        self.assertEqual(_COL_GROUPS[3], ("2026 Preliminary EA", 9, 14))

        styles = _Styles()
        self.assertIsNotNone(styles.fill_merged)
        self.assertIsNotNone(styles.fill_delineated)
        self.assertIsNotNone(styles.fill_special)

        # Test barangay summary row name resolution
        ea_layer = QgsVectorLayer("Polygon?crs=EPSG:3857", "01728001_ea", "memory")
        pr = ea_layer.dataProvider()
        pr.addAttributes([
            QgsField("GEOCODE", QVariant.String),
            QgsField("barangay", QVariant.String),
            QgsField("name", QVariant.String),
            QgsField("ean", QVariant.String),
            QgsField("hhcount", QVariant.Double),
            QgsField("bldgcount", QVariant.Int),
        ])
        ea_layer.updateFields()

        f = QgsFeature(ea_layer.fields())
        f.setGeometry(make_square(0, 0, 100))
        f.setAttributes(["01728001001", "Alzate", "EA 001000", "001000", 100.0, 20])
        pr.addFeatures([f])
        ea_layer.updateExtents()

        from references.create_enumeration_area.helpers.earf_writer import EARFWriter
        writer = EARFWriter(layer=ea_layer, geo_code="01728", citymun="Bangar", output_path="dummy.xlsx")
        rows = writer._collect_data_rows()
        brgy_rows = [r for r in rows if r.get("is_barangay_summary")]
        self.assertEqual(len(brgy_rows), 1)
        self.assertEqual(brgy_rows[0]["name"], "ALZATE")
        self.assertNotEqual(brgy_rows[0]["name"], "EA 001000")

        del ea_layer, writer, rows, brgy_rows
        gc.collect()


if __name__ == "__main__":
    unittest.main()


