# -*- coding: utf-8 -*-
"""
EARF Writer
-----------
Produces a styled **2026 Preliminary Enumeration Area Reference File (EARF)**
Excel workbook from the Tab 3 merged output layer.

Template layout (matching PSA census format):

  Row  1 : Title  — "2026 Preliminary Enumeration Area Reference File"
  Row  2 : Sub-title — "As of <Mmm> <YYYY>"
  Row  3 : (blank spacer)
  Row  4 : Column group headers (merged):
             Geographic Identification | 2024 EARF | 2024 Estimated |
             2026 Preliminary EAs
  Rows 5–8: Sub-header labels (merged vertically, word-wrapped)
  Row  9 : Column number codes  (1) (2) … (15)
  Row 10 : (blank spacer)
  Row 11+: Data block
             - City/Mun summary row  (bold, gray fill)
             - For each barangay:
                 Barangay summary row (bold, lighter fill)
                 EA data rows         (one per feature)

Column layout (A–O, 15 columns):
  A  Reg          — 2-digit region code
  B  Prov         — 5-digit province/city-mun code
  C  Mun          — city-mun portion
  D  Brgy         — 3-digit barangay suffix
  E  EA           — EA code (ean)
  F  Number of EAs                                         2024 EARF
  G  Province, City, Municipality, Barangay, and EA        2024 EARF
  H  Number of Households                                  2024 Estimated
  I  Number of Buildings                                   2024 Estimated
  J  New Enumeration Area Code                             2026 Preliminary EAs
  K  Number of Household                                   2026 Preliminary EAs
  L  Number of Buildings                                   2026 Preliminary EAs
  M  EA Type                                               2026 Preliminary EAs
  N  Source Year                                           2026 Preliminary EAs
  O  Remarks                                               2026 Preliminary EAs

Data sources (standard 19-field merged-layer schema):
  geocode (8-digit) → Reg / Prov / Mun / Brgy codes
  ean               → col E
  eacount           → col F
  name              → col G
  hhcount           → col H  (2024 baseline)
  bldgcount         → col I  (2024 baseline)
  new_ean           → col J
  hh_count          → col K  (2026 estimated)
  bldg_count        → col L  (2026 estimated)
  ea_type           → col M
  sy                → col N
  remarks           → col O

Delineated EAs : Each child feature appears on its own row.
Merged EAs     : Included once (prevailing EA); "Merged EA" appended to remarks.
Special EAs    : Included with ea_type as-is (GAP / OVERLAP / SPECIAL).

Dependencies: openpyxl >= 3.0
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import Any, Dict, List, Optional

# ---------------------------------------------------------------------------
# Column indices (0-based, A=0 … N=13)
# ---------------------------------------------------------------------------
_COL_PROV       = 0   # A
_COL_MUN        = 1   # B
_COL_BRGY       = 2   # C
_COL_EA         = 3   # D
_COL_EACOUNT    = 4   # E
_COL_NAME       = 5   # F
_COL_HHCOUNT    = 6   # G
_COL_BLDGCOUNT  = 7   # H
_COL_NEW_EAN    = 8   # I
_COL_HH_COUNT   = 9   # J
_COL_BLDG_COUNT = 10  # K
_COL_EA_TYPE    = 11  # L
_COL_SY         = 12  # M
_COL_REMARKS    = 13  # N

_TOTAL_COLS = 14   # A–N

# ---------------------------------------------------------------------------
# Row constants (1-indexed for openpyxl)
# ---------------------------------------------------------------------------
_ROW_TITLE     = 1
_ROW_AS_OF     = 2
_ROW_GRP_HDR   = 3
_ROW_SUBHDR    = 4
_ROW_NUM_CODES = 5
_DATA_START    = 6

# ---------------------------------------------------------------------------
# Column-group definitions  (1-indexed, inclusive)
# Format: (label, start_col, end_col)
# ---------------------------------------------------------------------------
_COL_GROUPS = [
    ("2024 EARF",           5,  6),   # E–F
    ("2024 Estimated",      7,  8),   # G–H
    ("2026 Preliminary EA", 9, 14),   # I–N
]

# ---------------------------------------------------------------------------
# Sub-header texts per column (displayed in row 4)
# ---------------------------------------------------------------------------
_SUBHDR = [
    "Prov",
    "Mun",
    "Brgy",
    "EA",
    "Number of\nEAs",
    "Province, City, Municipality,\nBarangay, and EA",
    "Number of\nHouseholds",
    "Number of\nBuildings",
    "New Enumeration\nArea Code",
    "Household\nCount",
    "Building\nCount",
    "EA Type",
    "Source\nYear",
    "Remarks",
]

# Number codes row 5
_NUM_CODES = [
    "(2)", "(3)", "(4)", "(5)",
    "(6)", "(7)", "(8)", "(9)",
    "(10)", "(11)", "(12)", "(13)", "(14)", "",
]

# Column widths (Excel character units)
_COL_WIDTHS = [
    7, 5, 6, 10,        # A–D (Prov, Mun, Brgy, EA)
    10, 44, 14, 12,     # E–H (Number of EAs, Name, HH, Bldg)
    18, 14, 14, 12, 10, 32,  # I–N (New EA Code … Remarks)
]

# Numeric column indices (0-based) — right-align
_NUMERIC_COLS = {
    _COL_EACOUNT, _COL_HHCOUNT, _COL_BLDGCOUNT,
    _COL_HH_COUNT, _COL_BLDG_COUNT,
}
# Center-aligned columns (codes, type, year)
_CENTER_COLS = {
    _COL_PROV, _COL_MUN, _COL_BRGY,
    _COL_EA, _COL_NEW_EAN, _COL_EA_TYPE, _COL_SY,
}


# ===========================================================================
# EARFWriter
# ===========================================================================

class EARFWriter:
    """Writes the 2026 Preliminary EARF Excel workbook.

    Parameters
    ----------
    layer : QgsVectorLayer
        The merged EA output layer (standard 19-field schema).
    geo_code : str
        5-digit geographic code (city/municipality level).
    citymun : str
        City/Municipality name used in the summary row.
    output_path : str
        Full path for the output .xlsx file.
    as_of_date : datetime, optional
        Date for the "As of <Mmm> <YYYY>" sub-title. Defaults to today.
    feedback : callable, optional
        ``callback(msg: str)`` for logging. Defaults to no-op.
    """

    def __init__(
        self,
        layer,
        geo_code: str,
        citymun: str,
        output_path: str,
        as_of_date: Optional[datetime] = None,
        feedback=None,
        ghost_features: Optional[list] = None,
    ) -> None:
        self._layer         = layer
        self._geo_code      = geo_code
        self._citymun       = citymun
        self._output_path   = output_path
        self._as_of_date    = as_of_date or datetime.now()
        self._log           = feedback or (lambda msg: None)
        # Fully-consumed previous EA features (ghost rows for Excel only).
        # These have NULL geometry and ea_type=MERGED; they are included in
        # the Excel output as reference rows but excluded from the .gpkg layer.
        self._ghost_features: list = ghost_features or []

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    def write(self) -> bool:
        """Generate the EARF Excel file. Returns True on success."""
        try:
            import openpyxl
        except ImportError:
            self._log("[ERROR] openpyxl is not installed. Cannot generate EARF Excel.")
            return False

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "EA2026"

            self._log("[INFO] EARF Writer: collecting data rows...")
            data_rows = self._collect_data_rows()

            styles = _Styles()
            self._log("[INFO] EARF Writer: writing title block...")
            self._write_title_block(ws, styles)

            self._log("[INFO] EARF Writer: writing column group headers...")
            self._write_group_headers(ws, styles)

            self._log("[INFO] EARF Writer: writing sub-headers...")
            self._write_sub_headers(ws, styles)

            self._log("[INFO] EARF Writer: writing data rows...")
            self._write_data_block(ws, styles, data_rows)

            # Column widths
            for col_idx, width in enumerate(_COL_WIDTHS, start=1):
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            # Row heights for header section
            ws.row_dimensions[_ROW_TITLE].height     = 18
            ws.row_dimensions[_ROW_AS_OF].height     = 15
            ws.row_dimensions[_ROW_GRP_HDR].height   = 18
            ws.row_dimensions[_ROW_SUBHDR].height    = 28
            ws.row_dimensions[_ROW_NUM_CODES].height = 14

            # Freeze panes: keep headers visible when scrolling
            ws.freeze_panes = ws.cell(row=_DATA_START, column=1)

            wb.save(self._output_path)
            self._log(f"[INFO] EARF Excel written: {self._output_path}")
            return True

        except Exception as exc:
            import traceback
            self._log(f"[ERROR] EARF Writer failed: {exc}")
            self._log(f"[ERROR] {traceback.format_exc()}")
            return False

    # ------------------------------------------------------------------
    # Data collection
    # ------------------------------------------------------------------

    def _collect_data_rows(self) -> List[Dict[str, Any]]:
        """Read features from the layer and return ordered row dicts.

        Order:
          1. Province summary row
          2. City/Mun summary row
          3. For each barangay (sorted by 8-digit geocode):
             a. Barangay summary row
             b. EA rows sorted by Mother EA and new_ean
        """
        try:
            from qgis.core import NULL
        except ImportError:
            NULL = None  # allow unit testing outside QGIS

        layer  = self._layer
        fields = layer.fields()
        name_to_idx: Dict[str, int] = {
            fields.at(i).name().lower(): i for i in range(fields.count())
        }

        def _str(feat, fname: str) -> Optional[str]:
            idx = name_to_idx.get(fname.lower(), -1)
            if idx == -1:
                return None
            val = feat.attribute(idx)
            if val is None or val == NULL:
                return None
            s = str(val).strip()
            return None if s in ("", "NULL", "None") else s

        def _remarks_str(feat, fname: str) -> str:
            idx = name_to_idx.get(fname.lower(), -1)
            if idx == -1:
                return ""
            val = feat.attribute(idx)
            if val is None or val == NULL:
                return ""
            s = str(val).strip()
            if s.upper() in ("", "NULL", "NONE", "FALSE", "0", "F"):
                return ""
            return s

        def _num(feat, fname: str) -> Optional[Any]:
            raw = _str(feat, fname)
            if raw is None:
                return None
            try:
                f = float(raw)
                return int(f) if f == int(f) else f
            except (ValueError, TypeError):
                return None

        # ── Collect raw EA rows ──────────────────────────────────────────
        raw_rows: List[Dict[str, Any]] = []
        prov_name = ""

        for feat in layer.getFeatures():
            geocode_raw = _str(feat, "geocode") or ""
            digits = re.sub(r"\D", "", geocode_raw)

            # Parse geographic code components (8-digit: PPPMMBBB)
            # Breakdown: digits[0:3]=prov, digits[3:5]=mun, digits[5:8]=brgy
            if len(digits) >= 8:
                prov = digits[0:3]    # 3-digit province code
                mun  = digits[3:5]    # 2-digit municipality code
                brgy = digits[5:8]    # 3-digit barangay suffix
                geocode_8 = digits[:8]
            elif len(digits) >= 5:
                prov = digits[0:3]
                mun  = digits[3:5]
                brgy = ""
                geocode_8 = digits[:5].ljust(8, "0")
            else:
                prov = mun = brgy = ""
                geocode_8 = (digits or geocode_raw).ljust(8, "0")[:8]

            ean_val     = _str(feat, "ean") or _str(feat, "ea_no") or _str(feat, "ea") or _str(feat, "eano") or ""
            new_ean_val = _str(feat, "new_ean") or _str(feat, "new_eacode") or ""
            remarks_val = _remarks_str(feat, "remarks")
            brgy_name_val = (
                _str(feat, "barangay")
                or _str(feat, "bgy_name")
                or _str(feat, "brgy_name")
                or _str(feat, "barangay_name")
                or _str(feat, "bgy_desc")
                or _str(feat, "brgy_desc")
                or _str(feat, "adm4_en")
                or _str(feat, "bgy")
                or _str(feat, "brgy")
                or ""
            )

            p_cand = _str(feat, "province") or _str(feat, "prov_name") or _str(feat, "prov") or ""
            if p_cand and not prov_name:
                prov_name = p_cand

            raw_rows.append({
                # Geographic identifiers
                "prov":          prov,
                "mun":           mun,
                "brgy":          brgy,
                "ea":            ean_val,
                "geocode_8":     geocode_8,
                "barangay_name": brgy_name_val,
                # 2024 EARF columns
                "eacount":    _num(feat, "eacount") if _num(feat, "eacount") is not None else 1,
                "name":       _str(feat, "name")     or "",
                "hhcount":    _num(feat, "hhcount"),
                "bldgcount":  _num(feat, "bldgcount"),
                # 2026 Preliminary columns
                "new_ean":    new_ean_val,
                "hh_count":   _num(feat, "hh_count"),
                "bldg_count": _num(feat, "bldg_count"),
                "ea_type":    _str(feat, "ea_type")  or "RETAINED",
                "sy":         _str(feat, "sy")        or "2026",
                "remarks":    remarks_val,
                # Row type
                "is_province_summary": False,
                "is_citymun_summary":  False,
                "is_barangay_summary": False,
                "is_ea_row":           True,
                "is_ghost":            False,
            })

        # ── Process ghost rows (fully-consumed previous EAs) ────────────
        active_keys = {
            (r.get("geocode_8", ""), r.get("ea", ""))
            for r in raw_rows
            if not r.get("is_ghost") and r.get("ea")
        }

        for ghost_feat in self._ghost_features:
            try:
                ghost_fields = ghost_feat.fields()
            except Exception:
                continue
            if ghost_fields is None:
                continue

            ghost_name_to_idx: Dict[str, int] = {
                ghost_fields.at(i).name().lower(): i
                for i in range(ghost_fields.count())
            }

            def _g_str(fname: str) -> Optional[str]:
                idx = ghost_name_to_idx.get(fname.lower(), -1)
                if idx == -1:
                    return None
                val = ghost_feat.attribute(idx)
                if val is None or val == NULL:
                    return None
                s = str(val).strip()
                return None if s in ("", "NULL", "None") else s

            def _g_remarks(fname: str) -> str:
                idx = ghost_name_to_idx.get(fname.lower(), -1)
                if idx == -1:
                    return ""
                val = ghost_feat.attribute(idx)
                if val is None or val == NULL:
                    return ""
                s = str(val).strip()
                if s.upper() in ("", "NULL", "NONE", "FALSE", "0", "F"):
                    return ""
                return s

            def _g_num(fname: str) -> Optional[Any]:
                raw = _g_str(fname)
                if raw is None:
                    return None
                try:
                    f = float(raw)
                    return int(f) if f == int(f) else f
                except (ValueError, TypeError):
                    return None

            geocode_raw = _g_str("geocode") or ""
            digits = re.sub(r"\D", "", geocode_raw)
            if len(digits) >= 8:
                prov = digits[0:3]
                mun  = digits[3:5]
                brgy = digits[5:8]
                geocode_8 = digits[:8]
            elif len(digits) >= 5:
                prov = digits[0:3]
                mun  = digits[3:5]
                brgy = ""
                geocode_8 = digits[:5].ljust(8, "0")
            else:
                prov = mun = brgy = ""
                geocode_8 = (digits or geocode_raw).ljust(8, "0")[:8]

            ghost_ea = _g_str("ean") or _g_str("ea_no") or _g_str("ea") or _g_str("eano") or ""
            if not ghost_ea or (geocode_8, ghost_ea) in active_keys:
                # Already represented by an active output feature; do not duplicate
                continue

            ghost_rmk = _g_remarks("remarks")
            if not ghost_rmk:
                ghost_rmk = "Merged EA"

            ghost_brgy_name = (
                _g_str("barangay")
                or _g_str("bgy_name")
                or _g_str("brgy_name")
                or _g_str("barangay_name")
                or _g_str("bgy_desc")
                or _g_str("brgy_desc")
                or _g_str("adm4_en")
                or _g_str("bgy")
                or _g_str("brgy")
                or ""
            )

            p_cand = _g_str("province") or _g_str("prov_name") or _g_str("prov") or ""
            if p_cand and not prov_name:
                prov_name = p_cand

            raw_rows.append({
                "prov":          prov,
                "mun":           mun,
                "brgy":          brgy,
                "ea":            ghost_ea,
                "geocode_8":     geocode_8,
                "barangay_name": ghost_brgy_name,
                # 2024 EARF columns — carry original counts for reference
                "eacount":    _g_num("eacount") or 1,
                "name":       _g_str("name")    or "",
                "hhcount":    _g_num("hhcount"),
                "bldgcount":  _g_num("bldgcount"),
                # 2026 Preliminary columns — left empty for merged partner
                "new_ean":    "",
                "hh_count":   None,
                "bldg_count": None,
                "ea_type":    _g_str("ea_type") or "MERGED",
                "sy":         _g_str("sy")       or "2024",
                "remarks":    ghost_rmk,
                # Row type
                "is_province_summary": False,
                "is_citymun_summary":  False,
                "is_barangay_summary": False,
                "is_ea_row":           True,
                "is_ghost":            True,
            })

        # ── City/Mun and Province totals ────────────────────────────────
        def _sum(rows, key):
            vals = [r[key] for r in rows if r.get(key) is not None]
            return sum(vals) if vals else None

        first = raw_rows[0] if raw_rows else {}
        first_prov = first.get("prov", "") or (self._geo_code[:3] if len(self._geo_code) >= 3 else "")
        first_mun = first.get("mun", "") or (self._geo_code[3:5] if len(self._geo_code) >= 5 else "")

        # Level 1: Province summary row (PPP 00 000 000000)
        province_row = {
            "prov":                first_prov,
            "mun":                 "00",
            "brgy":                "000",
            "ea":                  "000000",
            "geocode_8":           first_prov.ljust(8, "0"),
            "eacount":             _sum(raw_rows, "eacount"),
            "name":                prov_name.upper() if prov_name else f"PROVINCE {first_prov}",
            "hhcount":             _sum(raw_rows, "hhcount"),
            "bldgcount":           _sum(raw_rows, "bldgcount"),
            "new_ean":             "",
            "hh_count":            None,
            "bldg_count":          None,
            "ea_type":             "",
            "sy":                  "",
            "remarks":             "",
            "is_province_summary": True,
            "is_citymun_summary":  False,
            "is_barangay_summary": False,
            "is_ea_row":           False,
            "is_ghost":            False,
        }

        # Level 2: City/Mun summary row (PPP MM 000 000000)
        citymun_row = {
            "prov":                first_prov,
            "mun":                 first_mun,
            "brgy":                "000",
            "ea":                  "000000",
            "geocode_8":           self._geo_code.ljust(8, "0")[:8],
            "eacount":             None,
            "name":                self._citymun.upper(),
            "hhcount":             None,
            "bldgcount":           None,
            "new_ean":             "",
            "hh_count":            None,
            "bldg_count":          None,
            "ea_type":             "",
            "sy":                  "",
            "remarks":             "",
            "is_province_summary": False,
            "is_citymun_summary":  True,
            "is_barangay_summary": False,
            "is_ea_row":           False,
            "is_ghost":            False,
        }

        # ── Group by barangay ────────────────────────────────────────────
        brgy_groups: Dict[str, List[Dict]] = {}
        for r in raw_rows:
            brgy_groups.setdefault(r["geocode_8"], []).append(r)

        # ── Build ordered output ─────────────────────────────────────────
        output: List[Dict[str, Any]] = [province_row, citymun_row]

        for brgy_code in sorted(brgy_groups.keys()):
            ea_rows = brgy_groups[brgy_code]
            first_ea = ea_rows[0]

            # Derive barangay name:
            # 1. First check if any row has an explicit barangay name field
            brgy_name = ""
            for r in ea_rows:
                candidate = r.get("barangay_name", "").strip()
                if candidate and not re.match(r"^EA\s+\d+", candidate, re.IGNORECASE):
                    brgy_name = candidate.upper()
                    break

            # 2. If not found, derive from name field by stripping EA suffix
            if not brgy_name:
                for r in ea_rows:
                    cand = r.get("name", "").strip()
                    if cand:
                        cleaned = re.sub(
                            r"\s*[-\u2013]\s*EA\s+\S+\s*$", "", cand, flags=re.IGNORECASE
                        ).strip()
                        if cleaned and not re.match(r"^EA\s+\d+", cleaned, re.IGNORECASE):
                            brgy_name = cleaned.upper()
                            break

            # 3. Fallback to BARANGAY <suffix> (e.g. BARANGAY 001)
            if not brgy_name:
                brgy_suffix = first_ea.get("brgy", "") or brgy_code[-3:]
                brgy_name = f"BARANGAY {brgy_suffix}".upper()

            # Level 3: Barangay summary row
            brgy_row = {
                "prov":                first_ea.get("prov", ""),
                "mun":                 first_ea.get("mun",  ""),
                "brgy":                first_ea.get("brgy", ""),
                "ea":                  "000000",
                "geocode_8":           brgy_code,
                "eacount":             None,
                "name":                brgy_name,
                "hhcount":             None,
                "bldgcount":           None,
                "new_ean":             "",
                "hh_count":            None,
                "bldg_count":          None,
                "ea_type":             "",
                "sy":                  "",
                "remarks":             "",
                "is_province_summary": False,
                "is_citymun_summary":  False,
                "is_barangay_summary": True,
                "is_ea_row":           False,
                "is_ghost":            False,
            }
            output.append(brgy_row)

            # Enrich and guarantee complete remarks for all merged EAs in this barangay
            active_merged = [
                r for r in ea_rows
                if (r.get("ea_type") or "").upper() == "MERGED" and not r.get("is_ghost")
            ]
            ghost_merged = [
                r for r in ea_rows
                if (r.get("ea_type") or "").upper() == "MERGED" and r.get("is_ghost")
            ]

            if ghost_merged and active_merged:
                prevailing_ean = active_merged[0].get("new_ean") or active_merged[0].get("ea") or ""
                # Ensure each ghost partner row states the prevailing EA it merged into
                for g_row in ghost_merged:
                    cur_rmk = (g_row.get("remarks") or "").strip()
                    if not cur_rmk or cur_rmk.upper() in ("", "MERGED EA", "MERGED"):
                        g_row["remarks"] = f"Merged to EA {prevailing_ean}"
                    elif "merged to" not in cur_rmk.lower() and "merged with" not in cur_rmk.lower():
                        g_row["remarks"] = f"Merged to EA {prevailing_ean}; {cur_rmk}"

                # Ensure each active prevailing merged EA row lists its absorbed partner(s)
                for a_row in active_merged:
                    absorbed_list = [
                        g.get("ea") for g in ghost_merged
                        if g.get("ea") and g.get("ea") != a_row.get("ea") and g.get("ea") != a_row.get("new_ean")
                    ]
                    if absorbed_list:
                        cur_rmk = (a_row.get("remarks") or "").strip()
                        if not cur_rmk or cur_rmk.upper() in ("", "MERGED EA", "MERGED"):
                            a_row["remarks"] = f"Merged with EA {', '.join(absorbed_list)}"
                        elif "merged" not in cur_rmk.lower():
                            a_row["remarks"] = f"Merged with EA {', '.join(absorbed_list)}; {cur_rmk}"

            # Level 4: EA rows under this barangay grouped by Mother EA
            mother_ea_groups: Dict[str, List[Dict]] = {}
            for r in ea_rows:
                m_ea = r.get("ea", "") or ""
                mother_ea_groups.setdefault(m_ea, []).append(r)

            for m_ea in sorted(mother_ea_groups.keys()):
                group_rows = mother_ea_groups[m_ea]
                # Sort child rows by new_ean
                group_rows.sort(key=lambda r: (r.get("new_ean", "") or "", r.get("remarks", "") or ""))

                for idx, ea_row in enumerate(group_rows):
                    if idx > 0:
                        # For subsequent child rows of a delineated EA:
                        # Suppress 2024 baseline counts to prevent double-counting
                        ea_row["eacount"] = None
                        ea_row["hhcount"] = None
                        ea_row["bldgcount"] = None
                    output.append(ea_row)

        return output

    # ------------------------------------------------------------------
    # Section writers
    # ------------------------------------------------------------------

    def _write_title_block(self, ws, styles) -> None:
        """Rows 1–2: title and as-of date."""
        # Row 1
        cell = ws.cell(row=_ROW_TITLE, column=1,
                       value="2026 Preliminary Enumeration Area Reference File")
        cell.font      = styles.font_title
        cell.alignment = styles.align_left
        ws.merge_cells(
            start_row=_ROW_TITLE, start_column=1,
            end_row=_ROW_TITLE,   end_column=_TOTAL_COLS,
        )

        # Row 2
        as_of_str = self._as_of_date.strftime("As of %b %Y")
        cell2 = ws.cell(row=_ROW_AS_OF, column=1, value=as_of_str)
        cell2.font      = styles.font_normal_bold
        cell2.alignment = styles.align_left
        ws.merge_cells(
            start_row=_ROW_AS_OF, start_column=1,
            end_row=_ROW_AS_OF,   end_column=_TOTAL_COLS,
        )

    def _write_group_headers(self, ws, styles) -> None:
        """Row 3: merged column-group header cells."""
        group_fills = {
            "2024 EARF":           styles.fill_2024,
            "2024 Estimated":      styles.fill_est,
            "2026 Preliminary EA": styles.fill_2026,
        }

        # Apply borders to columns 1..4 in row 3
        for c in range(1, 5):
            mc = ws.cell(row=_ROW_GRP_HDR, column=c)
            mc.border = styles.border_thin

        for label, start_col, end_col in _COL_GROUPS:
            fill = group_fills.get(label, styles.fill_subhdr)

            # Write label in the first cell of the group
            cell = ws.cell(row=_ROW_GRP_HDR, column=start_col, value=label)
            cell.font      = styles.font_group_hdr
            cell.alignment = styles.align_center
            cell.border    = styles.border_thin

            # Merge if multi-column
            if start_col < end_col:
                ws.merge_cells(
                    start_row=_ROW_GRP_HDR, start_column=start_col,
                    end_row=_ROW_GRP_HDR,   end_column=end_col,
                )

            # Apply fill to every cell in the merged range
            for c in range(start_col, end_col + 1):
                mc = ws.cell(row=_ROW_GRP_HDR, column=c)
                mc.fill   = fill
                mc.border = styles.border_thin

    def _write_sub_headers(self, ws, styles) -> None:
        """Row 4 subheaders + Row 5 number codes."""
        # Row 4 — Sub-headers
        for col_idx, label in enumerate(_SUBHDR, start=1):
            cell = ws.cell(row=_ROW_SUBHDR, column=col_idx, value=label)
            cell.font      = styles.font_subhdr
            cell.alignment = styles.align_center_wrap
            cell.fill      = styles.fill_subhdr
            cell.border    = styles.border_thin

        # Row 5 — Number codes
        for col_idx, code in enumerate(_NUM_CODES, start=1):
            cell = ws.cell(row=_ROW_NUM_CODES, column=col_idx, value=code)
            cell.font      = styles.font_num_code
            cell.alignment = styles.align_center
            cell.fill      = styles.fill_subhdr
            cell.border    = styles.border_thin

    def _write_data_block(self, ws, styles,
                          data_rows: List[Dict[str, Any]]) -> None:
        """Write province summary, city/mun summary, barangay summaries, and EA rows."""
        for offset, row_data in enumerate(data_rows):
            row = _DATA_START + offset
            if row_data.get("is_province_summary"):
                self._write_row(ws, styles, row, row_data, kind="province")
            elif row_data.get("is_citymun_summary"):
                self._write_row(ws, styles, row, row_data, kind="citymun")
            elif row_data.get("is_barangay_summary"):
                self._write_row(ws, styles, row, row_data, kind="barangay")
            else:
                self._write_row(ws, styles, row, row_data, kind="ea")

    def _write_row(self, ws, styles, row: int,
                   row_data: Dict, kind: str) -> None:
        """Write a single row; kind = 'province' | 'citymun' | 'barangay' | 'ea'."""
        if kind == "province":
            font   = styles.font_summary_province
            fill   = styles.fill_summary_province
            border = styles.border_summary_province
        elif kind == "citymun":
            font   = styles.font_summary_citymun
            fill   = styles.fill_summary_citymun
            border = styles.border_thin
        elif kind == "barangay":
            font   = styles.font_summary_brgy
            fill   = styles.fill_summary_brgy
            border = styles.border_thin
        else:
            ea_type = (row_data.get("ea_type") or "").upper()
            is_ghost = row_data.get("is_ghost", False)

            if is_ghost or ea_type == "MERGED":
                font   = styles.font_ghost if is_ghost else styles.font_data
                fill   = styles.fill_merged
                border = styles.border_thin_light
            elif ea_type == "DELINEATED":
                font   = styles.font_data
                fill   = styles.fill_delineated
                border = styles.border_thin_light
            elif ea_type in ("SPECIAL", "GAP", "OVERLAP"):
                font   = styles.font_data
                fill   = styles.fill_special
                border = styles.border_thin_light
            else:
                font   = styles.font_data
                fill   = styles.fill_none
                border = styles.border_thin_light

        values = [
            row_data.get("prov",       ""),
            row_data.get("mun",        ""),
            row_data.get("brgy",       ""),
            row_data.get("ea",         ""),
            row_data.get("eacount"),
            row_data.get("name",       ""),
            row_data.get("hhcount"),
            row_data.get("bldgcount"),
            row_data.get("new_ean",    ""),
            row_data.get("hh_count"),
            row_data.get("bldg_count"),
            row_data.get("ea_type",    ""),
            row_data.get("sy",         ""),
            row_data.get("remarks",    ""),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font   = font
            cell.fill   = fill
            cell.border = border

            zero = col_idx - 1  # 0-based column index
            if zero in _NUMERIC_COLS:
                cell.alignment = styles.align_right
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0"
            elif zero in _CENTER_COLS:
                cell.alignment = styles.align_center
            else:
                cell.alignment = styles.align_left


# ===========================================================================
# _Styles — centralised openpyxl style objects
# ===========================================================================

class _Styles:
    """Create and cache all openpyxl style objects used by EARFWriter."""

    def __init__(self) -> None:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        # ── Fonts ────────────────────────────────────────────────────────
        self.font_title            = Font(name="Arial", size=12, bold=True)
        self.font_normal_bold      = Font(name="Arial", size=10, bold=True)
        self.font_group_hdr        = Font(name="Arial", size=9,  bold=True)
        self.font_subhdr           = Font(name="Arial", size=8,  bold=True)
        self.font_num_code         = Font(name="Arial", size=8,  bold=False)
        self.font_summary_province = Font(name="Arial", size=9,  bold=True)
        self.font_summary_citymun  = Font(name="Arial", size=9,  bold=True)
        self.font_summary_brgy     = Font(name="Arial", size=8,  bold=True)
        self.font_data             = Font(name="Arial", size=8)
        # Ghost rows: italic to signal they are reference-only rows
        self.font_ghost            = Font(name="Arial", size=8, italic=True)

        # ── Alignments ───────────────────────────────────────────────────
        self.align_left         = Alignment(horizontal="left",   vertical="center")
        self.align_center       = Alignment(horizontal="center", vertical="center")
        self.align_right        = Alignment(horizontal="right",  vertical="center")
        self.align_center_wrap  = Alignment(horizontal="center", vertical="center",
                                            wrap_text=True)

        # ── Fills ────────────────────────────────────────────────────────
        def _fill(hex_color: str) -> PatternFill:
            return PatternFill(fill_type="solid", fgColor=hex_color)

        self.fill_2024             = _fill("D6E4BC")   # 2024 EARF — light green
        self.fill_est              = _fill("FFFFC1")   # 2024 Estimated — light yellow
        self.fill_2026             = _fill("C5D9F1")   # 2026 Preliminary — light blue
        self.fill_subhdr           = _fill("F2F2F2")   # Sub-header — light gray
        self.fill_summary_province = _fill("E2EFDA")   # Province row — soft green
        self.fill_summary_citymun  = _fill("EDF2E8")   # City/Mun row — light soft green
        self.fill_summary_brgy     = _fill("EFEFEF")   # Barangay row — soft light gray

        # Row highlighting fills for EA categories:
        self.fill_merged          = _fill("FFF0E6")   # Merged EAs — Soft Light Peach
        self.fill_delineated      = _fill("FEF9E7")   # Delineated EAs — Soft Warm Butter
        self.fill_special         = _fill("ECEFF1")   # Special EAs — Soft Mist
        self.fill_ghost           = self.fill_merged
        self.fill_none            = PatternFill(fill_type=None)

        # ── Borders ──────────────────────────────────────────────────────
        _thin   = Side(style="thin")
        _double = Side(style="double")
        _hair   = Side(style="hair")

        self.border_thin = Border(
            left=_thin, right=_thin, top=_thin, bottom=_thin,
        )
        self.border_thin_light = Border(
            left=_hair, right=_hair, top=_hair, bottom=_hair,
        )
        self.border_summary_province = Border(
            left=_thin, right=_thin, top=_thin, bottom=_double,
        )
