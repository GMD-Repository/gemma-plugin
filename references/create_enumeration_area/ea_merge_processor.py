# -*- coding: utf-8 -*-
"""
EA Merge Processor
------------------
Implements the Enumeration Area Merge workflow for the EA Delineation and
Merging plugin (Tab 3).

Workflow:
  Phase 1 — Validate Previous EA Layer (geometry type, feature count, 5-digit
             geographic code, CityMun field).
  Phase 2 — Validate Replacement Polygon Layers (8-digit numeric names,
             polygon geometry, non-empty).
  Phase 3 — CRS reconciliation across all input layers.
  Phase 4 — Dissolve/union all replacement geometries into a single combined
             geometry (with spatial index).
  Phase 5 — For each EA feature, subtract the combined replacement geometry.
             Preserve original attributes for all remaining EA fragments.
  Phase 6 — Collect replacement features from all replacement layers.
  Phase 7 — Build output memory layer with the Previous EA Layer's field schema.
  Phase 8 — Validate final output geometries.
  Phase 9 — Add output layer to the QGIS project.
  Phase 10 — Export final attribute table to Excel.

Inputs:
  - ea_layer         : QgsVectorLayer (polygon) — the Previous EA Layer.
  - replacement_layers: list[QgsVectorLayer]    — replacement polygon layers
                        each named with exactly 8 numeric digits.

Outputs:
  - QgsVectorLayer in-memory named  <5digit>_ea2026
  - Excel file named  <5digit>_earf_<citymun>.xlsx

Source data protection:
  The Previous EA Layer and all replacement layers are NEVER modified or
  overwritten. All operations work on copies of geometries.

Dependencies:
  PyQGIS (QGIS 3.x LTR / PyQt5), openpyxl
"""

from __future__ import annotations

import os
import re
import tempfile
from dataclasses import dataclass, field
from typing import Callable, List, Optional

from qgis.core import (
    QgsCoordinateReferenceSystem,
    QgsCoordinateTransform,
    QgsFeature,
    QgsFeatureRequest,
    QgsField,
    QgsFields,
    QgsGeometry,
    QgsProject,
    QgsSpatialIndex,
    QgsVectorLayer,
    QgsWkbTypes,
)
from qgis.PyQt.QtCore import QVariant

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Fields searched (case-insensitive) for the 5-digit geographic code.
_GEOCODE_FIELDS = (
    "geocode", "geo_code", "psgc", "adm4_pcode", "adm_pcode",
    "brgy_code", "ean", "ea_code",
)

# Fields searched (case-insensitive) for the City/Municipality name.
_CITYMUN_FIELDS = (
    "citymun", "city_mun", "mun_name", "municipality",
    "city", "municipality_name", "municity", "cityname", "munname",
)

# Pattern for valid replacement layer names.
# Accepts bare 8-digit names (01728011) or names with an underscore suffix
# (01728011_delineated_ea2026, 01728001_merged_ea2026, 01728009_special_ea).
_REPLACEMENT_NAME_RE = re.compile(r"^\d{8}(_|$)")

# Output year suffix.
_OUTPUT_YEAR = "ea2026"

# Excel sheet name.
_EXCEL_SHEET_NAME = "EA2026"


# ---------------------------------------------------------------------------
# Result Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class EAMergeSummary:
    """Statistical summary produced by EAMergeProcessor.run()."""
    geographic_code: str = ""
    previous_ea_layer_name: str = ""
    replacement_layer_count: int = 0
    replacement_feature_count: int = 0
    modified_ea_count: int = 0
    final_ea_feature_count: int = 0
    output_layer_name: str = ""
    output_file_path: str = ""
    excel_file_path: str = ""
    excel_file_name: str = ""
    citymun_name: str = ""
    overall_status: str = "READY"   # READY | PASS | WARNING | ERROR
    excel_generated: bool = False

    @property
    def ea_input_layer_name(self) -> str:
        return self.previous_ea_layer_name

    @ea_input_layer_name.setter
    def ea_input_layer_name(self, val: str) -> None:
        self.previous_ea_layer_name = val


@dataclass
class EAMergeResult:
    """Returned by EAMergeProcessor.run()."""
    success: bool = False
    error_message: str = ""
    output_layer: Optional[QgsVectorLayer] = None
    summary: EAMergeSummary = field(default_factory=EAMergeSummary)
    log_lines: List[str] = field(default_factory=list)


# Field candidate aliases for attribute mapping in Tab 3.
# Ensures original counts (hhcount, bldgcount) and calculated counts (hh_count, bldg_count)
# strictly follow their respective field lineages without cross-contamination.
_FIELD_CANDIDATE_MAP = {
    "fid": ("fid", "id", "objectid", "gid", "feat_id"),
    "map_uuid": ("map_uuid", "uuid", "guid", "mapuuid"),
    "geocode": ("geocode", "geo_code", "psgc", "adm4_pcode", "adm_pcode", "brgy_code"),
    "region": ("region", "reg_code", "reg_name", "adm1_pcode"),
    "province": ("province", "prov_code", "prov_name", "adm2_pcode"),
    "city_mun": ("city_mun", "citymun", "city_name", "mun_name", "municipality", "city", "adm3_pcode"),
    "barangay": ("barangay", "bgy_name", "brgy_name", "brgy", "bgy", "adm4_pcode"),
    "ean": ("ean", "ea_code", "ean_code", "ea_no", "eacode", "eano", "old_ean"),
    "name": ("name", "ean_name", "ea_name", "areaname", "area_name"),
    "code": ("code", "bgy_code", "brgy_code", "bgy_c", "brgy_c"),
    "hhcount": (
        "hhcount", "original_hhcount", "orig_hhcount", "orig_hh",
        "new_hhcount", "household", "household_count", "pop", "population"
    ),
    "bldgcount": (
        "bldgcount", "original_bldgcount", "orig_bldgcount", "orig_bldg",
        "new_bldgcount", "bldgpts_cnt", "bldg_points", "building_count",
        "bldg_total", "buildings"
    ),
    "sy": ("sy", "survey_yr", "survey_year", "year"),
    "new_ean": ("new_ean", "new_eacode", "new_ea", "ean_new", "new_ea_code", "new_ea_tracker"),
    "hh_count": (
        "hh_count", "new_hh_count", "calc_hh_count", "hh_cnt", "total_hh", "hh"
    ),
    "bldg_count": (
        "bldg_count", "new_bldg_count", "calc_bldg_count", "bldg_cnt", "bldg"
    ),
    "ea_type": ("ea_type", "type", "eatype", "special_type"),
    "eacount": ("eacount", "ea_count", "eacnt", "total_ea", "ea_total"),
    "remarks": ("remarks", "remark", "delin_remarks", "delin_remark", "comments", "comment"),
}

# The exact 19 output fields in strict order
_OUTPUT_FIELD_SPECS = (
    ("fid", QVariant.Int),
    ("map_uuid", QVariant.String),
    ("geocode", QVariant.String),
    ("region", QVariant.String),
    ("province", QVariant.String),
    ("city_mun", QVariant.String),
    ("barangay", QVariant.String),
    ("ean", QVariant.String),
    ("name", QVariant.String),
    ("code", QVariant.String),
    ("hhcount", QVariant.Double),
    ("bldgcount", QVariant.Int),
    ("sy", QVariant.String),
    ("new_ean", QVariant.String),
    ("hh_count", QVariant.Double),
    ("bldg_count", QVariant.Int),
    ("ea_type", QVariant.String),
    ("eacount", QVariant.Int),
    ("remarks", QVariant.String),
)


def _ea_type_from_layer_name(layer_name: str) -> str:
    """Derive EA type from layer name suffix (e.g. 01728011_delineated_ea2026 -> DELINEATED)."""
    name_lower = layer_name.lower()
    if "delineated" in name_lower:
        return "DELINEATED"
    if "merged" in name_lower:
        return "MERGED"
    if "special" in name_lower:
        return "SPECIAL"
    if "gap" in name_lower:
        return "GAP"
    if "overlap" in name_lower:
        return "OVERLAP"
    return "RETAINED"


def _extract_feature_attribute(
    feat: QgsFeature,
    name_to_idx: dict,
    target_field_name: str,
):
    """Extract an attribute value from feat for target_field_name using exact match
    or candidate alias matching."""
    name_lower = target_field_name.lower()
    from qgis.core import NULL

    def _is_valid(val):
        if val is None or val == NULL:
            return False
        s = str(val).strip()
        if s.upper() in ("", "NULL", "NONE", "FALSE", "F"):
            return False
        return True

    # 1. Try exact match
    exact_idx = name_to_idx.get(name_lower, -1)
    if exact_idx != -1:
        val = feat.attribute(exact_idx)
        if _is_valid(val):
            return val

    # 2. Try candidate aliases
    candidates = _FIELD_CANDIDATE_MAP.get(name_lower, ())
    for cand in candidates:
        cand_idx = name_to_idx.get(cand.lower(), -1)
        if cand_idx != -1:
            val = feat.attribute(cand_idx)
            if _is_valid(val):
                return val

    return None


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def _field_index_ci(layer: QgsVectorLayer, candidates: tuple) -> int:
    """Return the field index (0-based) for the first matching candidate name
    (case-insensitive). Returns -1 if not found."""
    fields = layer.fields()
    name_to_idx = {fields.at(i).name().lower(): i for i in range(fields.count())}
    for cand in candidates:
        idx = name_to_idx.get(cand.lower(), -1)
        if idx != -1:
            return idx
    return -1


def _first_nonempty_value(layer: QgsVectorLayer, field_idx: int) -> Optional[str]:
    """Return the first non-null, non-empty attribute value at field_idx."""
    if field_idx == -1:
        return None
    for feat in layer.getFeatures(QgsFeatureRequest().setFlags(
            QgsFeatureRequest.NoGeometry)):
        val = feat.attribute(field_idx)
        if val is None or val == "":
            continue
        from qgis.core import NULL
        if val == NULL:
            continue
        return str(val).strip()
    return None


def _unique_values(layer: QgsVectorLayer, field_idx: int) -> list:
    """Return a list of unique, non-null, non-empty attribute values at field_idx."""
    if field_idx == -1:
        return []
    seen = set()
    result = []
    from qgis.core import NULL
    for feat in layer.getFeatures(QgsFeatureRequest().setFlags(
            QgsFeatureRequest.NoGeometry)):
        val = feat.attribute(field_idx)
        if val is None or val == "" or val == NULL:
            continue
        s = str(val).strip()
        if s and s not in seen:
            seen.add(s)
            result.append(s)
    return result


def _repair_geometry(geom: QgsGeometry) -> QgsGeometry:
    """Return a repaired copy of the geometry if invalid; else return as-is."""
    if geom is None or geom.isNull() or geom.isEmpty():
        return geom
    if not geom.isGeosValid():
        fixed = geom.makeValid()
        if fixed and not fixed.isNull():
            return fixed
    return geom


# ---------------------------------------------------------------------------
# Main Processor Class
# ---------------------------------------------------------------------------

class EAMergeProcessor:
    """Executes the Enumeration Area Merge workflow."""

    @staticmethod
    def short_help_string() -> str:
        """Returns the HTML description string for Enumeration Area Merge."""
        return (
            "<h3>Enumeration Area Merge</h3>"
            "<p>The <b>Enumeration Area Merge</b> workflow updates an existing previous Enumeration Area (EA) "
            "layer using one or more replacement polygon layers containing updated/replacement EA geometries.</p>"
            "<p>Replacement polygons take precedence: any overlapping portions of the existing previous EA layer "
            "underneath the replacement geometries are removed, and the replacement geometries are inserted to "
            "produce a consolidated <code>&lt;5-digit geocode&gt;_ea2026</code> layer and an exact Excel attribute "
            "table export (<code>&lt;5-digit geocode&gt;_earf_&lt;citymun&gt;.xlsx</code>).</p>"

            "<h4>Inputs</h4>"
            "<b>Required</b>"
            "<ul>"
            "<li><b>Previous EA Layer</b> (polygon) — Starting/existing EA layer. Attributes and fields are preserved. "
            "Auto-detected by <code>*_ea</code>, <code>*_ea2024</code>, <code>*_ea2026</code>, or <code>*_ea_preprocessed</code>.</li>"
            "<li><b>Replacement Polygon Layers</b> (polygon, multi-input) — One or more replacement polygon layers from "
            "the project. Each layer name must follow the <b>8-digit numeric convention</b> (e.g. <code>01001000</code>).</li>"
            "</ul>"

            "<h4>Workflow & Guarantees</h4>"
            "<ul>"
            "<li><b>Source Data Protection</b>: Previous EA and replacement layers are never modified or overwritten.</li>"
            "<li><b>Precedence & Geometry Difference</b>: Overlapping areas in previous EAs are subtracted so replacement polygons take precedence without geometric overlap.</li>"
            "<li><b>Automated Geocode & CityMun Extraction</b>: Automatically determines the 5-digit geographic code and single City/Municipality name from the Previous EA layer.</li>"
            "<li><b>Consolidated Output Layer</b>: Adds <code>&lt;5-digit geocode&gt;_ea2026</code> directly to the active QGIS project.</li>"
            "<li><b>Excel Attribute Table Export</b>: Generates <code>&lt;5-digit geocode&gt;_earf_&lt;citymun&gt;.xlsx</code> with primary sheet <code>EA2026</code> mirroring output attribute columns.</li>"
            "</ul>"
        )

    def __init__(
        self,
        ea_layer: QgsVectorLayer,
        replacement_layers: List[QgsVectorLayer],
        output_dir: str = "",
        feedback_callback: Optional[Callable[[str], None]] = None,
        progress_callback: Optional[Callable[[int], None]] = None,
        is_cancelled_fn: Optional[Callable[[], bool]] = None,
        skip_add_to_project: bool = False,
    ):
        self.ea_layer = ea_layer
        self.replacement_layers = replacement_layers
        self.output_dir = output_dir or ""
        self._feedback = feedback_callback or (lambda msg: None)
        self._progress = progress_callback or (lambda pct: None)
        self._is_cancelled = is_cancelled_fn or (lambda: False)
        # When True the caller is responsible for adding the output layer to the
        # QGIS project (required when running inside a QThread because
        # QgsProject.addMapLayer() must be called from the main GUI thread).
        self._skip_add_to_project = skip_add_to_project

        # Populated during run()
        self._geo_code: str = ""
        self._citymun: str = ""
        self._output_layer_name: str = ""
        self._result = EAMergeResult()

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    def run(self) -> EAMergeResult:
        """Execute the full merge workflow and return an EAMergeResult."""
        self._result = EAMergeResult()
        summary = self._result.summary

        try:
            self._log("[INFO] Starting Enumeration Area Merge...")
            self._progress(2)

            # ── Phase 1: Validate Previous EA Layer ──────────────────────────
            self._log(f"[INFO] Previous EA Layer: {self.ea_layer.name()}")
            summary.previous_ea_layer_name = self.ea_layer.name()

            if not self._validate_ea_layer():
                return self._result

            # ── Phase 2: Determine 5-digit geographic code ─────────────────
            geo_code = self._determine_geographic_code()
            if geo_code is None:
                self._fail(
                    "Unable to determine the 5-digit geographic code from "
                    "the Previous EA Layer."
                )
                return self._result
            self._geo_code = geo_code
            summary.geographic_code = geo_code
            self._log(f"[INFO] Geographic Code: {geo_code}")
            self._progress(8)

            # ── Phase 3: Validate replacement layers ───────────────────────
            self._log(
                f"[INFO] Replacement Polygon Layers: "
                f"{len(self.replacement_layers)}"
            )
            summary.replacement_layer_count = len(self.replacement_layers)

            if not self._validate_replacement_layers():
                return self._result
            self._progress(15)

            if self._is_cancelled():
                self._fail("Cancelled by user.")
                return self._result

            # ── Phase 4: Validate geometries ───────────────────────────────
            self._log("[INFO] Validating geometries...")
            if not self._validate_geometries():
                return self._result
            self._progress(20)

            # ── Phase 5: Determine CityMun ─────────────────────────────────
            citymun = self._determine_citymun(self.ea_layer)
            if citymun is None:
                self._fail(
                    "Unable to determine the City/Municipality name required "
                    "for the Excel filename."
                )
                return self._result
            self._citymun = citymun
            self._log(f"[INFO] City/Municipality: {citymun}")
            self._progress(25)

            # ── Phase 6: Combine replacement geometries ────────────────────
            self._log("[INFO] Combining replacement geometries...")
            out_fields = self._build_output_fields()
            replacement_features, combined_geom, repl_feat_count = (
                self._prepare_replacement_geometries(out_fields)
            )
            if combined_geom is None or combined_geom.isNull():
                self._fail("Failed to combine replacement geometries.")
                return self._result
            summary.replacement_feature_count = repl_feat_count
            self._log(
                f"[INFO] Combined {repl_feat_count} replacement feature(s) "
                f"from {len(self.replacement_layers)} layer(s)."
            )
            self._progress(40)

            if self._is_cancelled():
                self._fail("Cancelled by user.")
                return self._result

            # ── Phase 7: Remove replacement areas from existing EAs ─────────
            self._log(
                "[INFO] Removing replacement areas from existing EA "
                "geometries..."
            )
            remaining_ea_features, ghost_ea_features, modified_count = (
                self._replace_ea_geometries(
                    combined_geom, out_fields, replacement_features
                )
            )
            summary.modified_ea_count = modified_count
            self._log(
                f"[INFO] Modified EA Features (geometry affected): "
                f"{modified_count}"
            )
            if ghost_ea_features:
                self._log(
                    f"[INFO] Fully-consumed EA features (ghost rows for "
                    f"Excel): {len(ghost_ea_features)}"
                )
            self._progress(65)

            if self._is_cancelled():
                self._fail("Cancelled by user.")
                return self._result

            # ── Phase 8: Combine into final feature list ───────────────────
            # ghost_ea_features are intentionally excluded here — they have
            # NULL geometry and must not enter the spatial (.gpkg) output.
            self._log("[INFO] Applying replacement geometries...")
            all_features = remaining_ea_features + replacement_features
            self._populate_ea_counts(all_features, out_fields)
            self._progress(70)

            # ── Phase 9: Create output memory layer ────────────────────────
            self._output_layer_name = f"{geo_code}_{_OUTPUT_YEAR}"
            self._log(
                f"[INFO] Creating output layer: {self._output_layer_name}"
            )
            output_layer = self._create_output_layer(all_features, out_fields)
            if output_layer is None:
                self._fail("Failed to create output layer.")
                return self._result
            summary.output_layer_name = self._output_layer_name
            summary.final_ea_feature_count = output_layer.featureCount()
            self._log(
                f"[INFO] Final EA features: "
                f"{output_layer.featureCount()}"
            )
            self._progress(80)

            # ── Phase 10: Validate final EA geometry ───────────────────────
            self._log("[INFO] Validating final EA geometry...")
            self._validate_output(output_layer)
            self._progress(85)

            # ── Phase 11: Export to GeoPackage and Add to QGIS project ─────
            saved_to_gpkg = False

            if output_layer.featureCount() == 0:
                self._log(
                    f"[INFO] Output layer '{self._output_layer_name}' has 0 features; "
                    "skipping adding layer to QGIS project."
                )
                self._result.output_layer = None
            else:
                if self.output_dir:
                    output_file_path = os.path.join(self.output_dir, f"{self._output_layer_name}.gpkg")
                else:
                    proj = QgsProject.instance() if QgsProject.instance() else None
                    proj_fn = proj.fileName() if proj and hasattr(proj, 'fileName') else None
                    proj_hp = proj.homePath() if proj and hasattr(proj, 'homePath') else None
                    if isinstance(proj_fn, str) and proj_fn.strip():
                        output_file_path = os.path.join(os.path.dirname(proj_fn), f"{self._output_layer_name}.gpkg")
                    elif isinstance(proj_hp, str) and proj_hp.strip():
                        output_file_path = os.path.join(proj_hp, f"{self._output_layer_name}.gpkg")
                    else:
                        output_file_path = os.path.join(tempfile.gettempdir(), f"{self._output_layer_name}.gpkg")

                from .helpers.pre_ea_detector import get_unique_filepath
                out_dir = os.path.dirname(output_file_path)
                base_fn = os.path.splitext(os.path.basename(output_file_path))[0]
                if not out_dir:
                    out_dir = tempfile.gettempdir()
                output_file_path = get_unique_filepath(out_dir, base_fn, ".gpkg")
                final_layer_name = os.path.splitext(os.path.basename(output_file_path))[0]

                try:
                    os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
                    if self._export_layer_to_gpkg(output_layer, output_file_path, final_layer_name):
                        self._log(f"[INFO] Output layer successfully saved to GeoPackage (.gpkg): {output_file_path}")
                        gpkg_layer = QgsVectorLayer(f"{output_file_path}|layername={final_layer_name}", final_layer_name, "ogr")
                        if not gpkg_layer.isValid():
                            gpkg_layer = QgsVectorLayer(output_file_path, final_layer_name, "ogr")

                        if gpkg_layer.isValid():
                            from .helpers.style import apply_qml_to_layer
                            apply_qml_to_layer(gpkg_layer, "ea_output.qml")
                            output_layer = gpkg_layer
                            self._output_layer_name = final_layer_name
                            summary.output_layer_name = final_layer_name
                            summary.output_file_path = output_file_path
                            saved_to_gpkg = True
                        else:
                            self._log(f"[ERROR] Could not load saved GeoPackage layer from: {output_file_path}")
                    else:
                        self._log(f"[ERROR] Failed to save GeoPackage (.gpkg) to {output_file_path}")
                except Exception as save_err:
                    self._log(f"[ERROR] Output layer GeoPackage export failed with exception: {save_err}")

                if self._skip_add_to_project:
                    # Caller (e.g. threaded dialog) will add the layer on the main
                    # thread once the worker finishes.
                    self._result.output_layer = output_layer
                    self._log(
                        f"[INFO] Output layer '{self._output_layer_name}' ready; "
                        "will be added to QGIS project by the calling thread."
                    )
                else:
                    proj = QgsProject.instance()
                    if proj and hasattr(proj, 'mapLayersByName'):
                        old_layers = proj.mapLayersByName(self._output_layer_name)
                        if old_layers and isinstance(old_layers, (list, tuple)):
                            for old_lyr in list(old_layers):
                                if hasattr(proj, 'removeMapLayer') and hasattr(old_lyr, 'id'):
                                    proj.removeMapLayer(old_lyr.id())
                        if hasattr(proj, 'addMapLayer'):
                            proj.addMapLayer(output_layer)
                    if saved_to_gpkg:
                        self._log(
                            f"[INFO] Permanent GeoPackage layer (.gpkg) added to QGIS canvas: {self._output_layer_name}"
                        )
                    else:
                        self._log(
                            f"[WARNING] Falling back to temporary in-memory layer: {self._output_layer_name}"
                        )
                    self._result.output_layer = output_layer
            self._progress(90)

            # ── Phase 12: Read final attribute table for Excel ─────────────
            self._log("[INFO] Reading final output attribute table...")

            # ── Phase 13: Export Preliminary EARF Excel ────────────────────
            excel_base = f"{geo_code}_earf_{citymun}"
            if self.output_dir:
                excel_path = get_unique_filepath(self.output_dir, excel_base, ".xlsx")
            else:
                # Default: same directory as the QGIS project file
                project_home = QgsProject.instance().homePath()
                if project_home:
                    excel_path = get_unique_filepath(project_home, excel_base, ".xlsx")
                else:
                    excel_path = get_unique_filepath(os.path.expanduser("~"), excel_base, ".xlsx")
            excel_name = os.path.basename(excel_path)

            self._log(
                f"[INFO] Generating Preliminary EARF Excel: {excel_name}"
            )

            # Primary: styled PSA EARF template via EARFWriter
            excel_ok = False
            try:
                from .helpers.earf_writer import EARFWriter
                writer = EARFWriter(
                    layer=output_layer,
                    geo_code=geo_code,
                    citymun=citymun,
                    output_path=excel_path,
                    feedback=self._feedback,
                    ghost_features=ghost_ea_features,
                )
                excel_ok = writer.write()
            except Exception as _earf_err:
                self._log(
                    f"[WARNING] EARFWriter raised an exception ({_earf_err}); "
                    "falling back to plain attribute table export."
                )

            # Fallback: plain attribute table dump (no styling)
            if not excel_ok:
                excel_ok = self._export_attribute_table_to_excel(
                    output_layer, excel_path,
                    ghost_features=ghost_ea_features,
                )

            if excel_ok:
                summary.excel_generated = True
                summary.excel_file_path = excel_path
                summary.excel_file_name = excel_name
                self._log(
                    "[INFO] Preliminary EARF Excel successfully generated."
                )
            else:
                # Warn but do not fail — the polygon output was created
                self._log(
                    f"[WARNING] The final EA layer was successfully created: "
                    f"{self._output_layer_name}\n"
                    "However, the Preliminary EARF Excel could not be generated."
                )

            self._progress(100)
            self._log("[INFO] Enumeration Area Merge completed.")
            summary.overall_status = "PASS"
            self._result.success = True

        except Exception as exc:
            import traceback
            tb = traceback.format_exc()
            self._log(
                f"[ERROR] Unexpected error during Enumeration Area Merge: "
                f"{exc}"
            )
            self._log(f"[ERROR] Traceback:\n{tb}")
            self._fail(str(exc))

        return self._result

    # ------------------------------------------------------------------
    # Phase helpers
    # ------------------------------------------------------------------

    def _validate_ea_layer(self) -> bool:
        """Validate the Previous EA Layer (geometry type, features)."""
        if self.ea_layer is None:
            self._fail("Previous EA Layer is required.")
            return False
        if self.ea_layer.geometryType() != QgsWkbTypes.PolygonGeometry:
            self._fail(
                f"Layer \"{self.ea_layer.name()}\" is not a polygon layer. "
                "A polygon layer is required for Previous EA Layer."
            )
            return False
        if self.ea_layer.featureCount() == 0:
            self._fail(
                f"Previous EA Layer \"{self.ea_layer.name()}\" contains no features."
            )
            return False
        self._log(
            f"[INFO] Previous EA Layer valid: "
            f"{self.ea_layer.featureCount()} polygons  "
            f"({self.ea_layer.crs().authid()})"
        )
        return True

    def _determine_geographic_code(self) -> Optional[str]:
        """Extract the 5-digit geographic code from the Previous EA Layer.

        Searches the first matching geocode field, then takes the first
        5 digits of the first non-null value.
        """
        geo_idx = _field_index_ci(self.ea_layer, _GEOCODE_FIELDS)
        raw = _first_nonempty_value(self.ea_layer, geo_idx)
        if not raw:
            return None

        # Strip to leading digits
        digits_only = re.sub(r"\D", "", raw)
        if len(digits_only) < 5:
            return None
        return digits_only[:5]

    def _validate_replacement_layers(self) -> bool:
        """Validate each replacement layer: polygon type, 8-digit name, non-empty."""
        if not self.replacement_layers:
            self._fail("At least one Replacement Polygon Layer is required.")
            return False

        self._log("[INFO] Validating replacement layer names...")

        for layer in self.replacement_layers:
            name = layer.name()

            # 1. 8-digit numeric name (prefix match — suffix after _ is allowed)
            if not _REPLACEMENT_NAME_RE.match(name):
                self._fail(
                    f"Layer \"{name}\" does not follow the required "
                    "8-digit numeric naming convention.\n"
                    "The replacement polygon layer name must begin with "
                    "exactly 8 numeric digits.\n"
                    "Accepted formats: ######## or ########_suffix "
                    "(e.g. 01728011_delineated_ea2026)"
                )
                return False

            # 2. Polygon geometry
            if layer.geometryType() != QgsWkbTypes.PolygonGeometry:
                self._fail(
                    f"Layer \"{name}\" is not a polygon layer. "
                    "Polygon layers are required for Enumeration Area Merge."
                )
                return False

            # 3. Non-empty
            if layer.featureCount() == 0:
                self._fail(
                    f"Replacement layer \"{name}\" contains no features."
                )
                return False

        self._log("[INFO] All replacement layers have valid 8-digit codes.")
        return True

    def _validate_geometries(self) -> bool:
        """Check EA and replacement layers for empty or null geometries.

        Invalid geometries are repaired on processing copies.
        Source layers are never modified.
        """
        # Check EA layer CRS is defined
        if not self.ea_layer.crs().isValid():
            self._fail(
                "Previous EA Layer has an undefined or invalid CRS. "
                "Please assign a valid CRS before processing."
            )
            return False

        # Check replacement layer CRSes are defined
        for layer in self.replacement_layers:
            if not layer.crs().isValid():
                self._fail(
                    f"Replacement layer \"{layer.name()}\" has an invalid CRS."
                )
                return False
        return True

    def _determine_citymun(self, layer: QgsVectorLayer) -> Optional[str]:
        """Return the single City/Municipality name from the layer attribute table.

        Returns None and logs an error if not found or if multiple distinct
        values exist (ambiguous).
        """
        citymun_idx = _field_index_ci(layer, _CITYMUN_FIELDS)
        if citymun_idx == -1:
            self._log(
                "[WARNING] No City/Municipality field found. "
                "Excel filename will omit the CityMun component."
            )
            return None

        values = _unique_values(layer, citymun_idx)
        if not values:
            return None
        if len(values) > 1:
            self._fail(
                "Multiple City/Municipality values found in the Previous EA Layer "
                f"({', '.join(values)}). "
                "Expected exactly one value for the Excel filename."
            )
            return None
        return values[0]

    def _build_output_fields(self) -> QgsFields:
        """Construct the output layer field schema.

        Constructs exactly the 19 standard output fields in the required order:
        1. fid (Int)
        2. map_uuid (String)
        3. geocode (String)
        4. region (String)
        5. province (String)
        6. city_mun (String)
        7. barangay (String)
        8. ean (String)
        9. name (String)
        10. code (String)
        11. hhcount (Double)
        12. bldgcount (Int)
        13. sy (String)
        14. new_ean (String)
        15. hh_count (Double)
        16. bldg_count (Int)
        17. ea_type (String)
        18. eacount (Int)
        19. remarks (String)

        Preserves existing field types from the Previous EA Layer if present.
        """
        ea_fields = self.ea_layer.fields() if self.ea_layer else None
        ea_field_map = {}
        if ea_fields:
            for i in range(ea_fields.count()):
                f = ea_fields.at(i)
                ea_field_map[f.name().lower()] = f

        out_fields = QgsFields()
        for fname, default_type in _OUTPUT_FIELD_SPECS:
            existing_field = ea_field_map.get(fname.lower())
            if existing_field is not None:
                out_fields.append(QgsField(fname, existing_field.type()))
            else:
                out_fields.append(QgsField(fname, default_type))

        return out_fields

    def _prepare_replacement_geometries(
        self,
        out_fields: Optional[QgsFields] = None,
    ) -> tuple[list[QgsFeature], Optional[QgsGeometry], int]:
        """Collect replacement features and produce a single dissolved union geometry.

        Returns:
          (replacement_features, combined_union_geometry, total_feature_count)

        CRS: all replacement geometries are transformed to the EA input CRS.
        """
        if out_fields is None:
            out_fields = self._build_output_fields()
        ea_crs = self.ea_layer.crs()
        replacement_features: list[QgsFeature] = []
        # Collect individual geometries first — build union in one shot at the
        # end via unaryUnion(), which is O(n log n) vs O(n²) incremental combine.
        geom_list: list[QgsGeometry] = []
        total_count = 0

        # Build spatial index on EA layer for attribute inheritance fallback
        ea_index = QgsSpatialIndex(self.ea_layer.getFeatures())
        ea_features_by_id = {f.id(): f for f in self.ea_layer.getFeatures()}
        ea_fields = self.ea_layer.fields()
        ea_name_to_idx = {
            ea_fields.at(i).name().lower(): i
            for i in range(ea_fields.count())
        }

        for layer in self.replacement_layers:
            layer_crs = layer.crs()
            needs_transform = (layer_crs != ea_crs)
            transform = None
            if needs_transform:
                transform = QgsCoordinateTransform(
                    layer_crs, ea_crs, QgsProject.instance()
                )
                self._log(
                    f"[INFO] CRS mismatch for \"{layer.name()}\": "
                    f"transforming from {layer_crs.authid()} "
                    f"to {ea_crs.authid()}."
                )

            repl_fields = layer.fields()
            # Build a case-insensitive name → index map for the replacement layer's
            # fields so we can look up values efficiently for every feature.
            repl_name_to_idx = {
                repl_fields.at(i).name().lower(): i
                for i in range(repl_fields.count())
            }

            for feat in layer.getFeatures():
                geom = QgsGeometry(feat.geometry())
                if geom is None or geom.isNull() or geom.isEmpty():
                    continue

                # Transform if needed
                if transform is not None:
                    geom.transform(transform)

                # Repair
                geom = _repair_geometry(geom)
                if geom is None or geom.isNull() or geom.isEmpty():
                    continue

                geom_list.append(QgsGeometry(geom))

                # Build output feature using the output fields schema.
                # Attributes follow specific field lineages:
                # - hhcount & bldgcount follow previous / replacement hhcount & bldgcount
                # - hh_count & bldg_count follow previous / replacement hh_count & bldg_count
                out_feat = QgsFeature(out_fields)
                out_feat.setGeometry(geom)

                # Find intersecting previous EA feature (if any) as fallback for missing attributes
                fallback_ea_feat = None
                intersecting_ids = ea_index.intersects(geom.boundingBox())
                best_overlap_area = 0.0
                for ea_fid in intersecting_ids:
                    prev_f = ea_features_by_id.get(ea_fid)
                    if prev_f and prev_f.geometry():
                        inter = geom.intersection(prev_f.geometry())
                        if inter and not inter.isEmpty():
                            area = inter.area()
                            if area > best_overlap_area:
                                best_overlap_area = area
                                fallback_ea_feat = prev_f

                for ea_i in range(out_fields.count()):
                    ea_field = out_fields.at(ea_i)
                    ea_field_name = ea_field.name().lower()
                    val = _extract_feature_attribute(feat, repl_name_to_idx, ea_field_name)

                    if val is None and fallback_ea_feat is not None:
                        # Fallback: inherit from overlapping previous EA feature
                        val = _extract_feature_attribute(fallback_ea_feat, ea_name_to_idx, ea_field_name)

                    # Secondary fallback if count / ea_type field was not present
                    if val is None:
                        if ea_field_name == "hh_count":
                            val = _extract_feature_attribute(feat, repl_name_to_idx, "hhcount")
                            if val is None and fallback_ea_feat is not None:
                                val = _extract_feature_attribute(fallback_ea_feat, ea_name_to_idx, "hhcount")
                        elif ea_field_name == "bldg_count":
                            val = _extract_feature_attribute(feat, repl_name_to_idx, "bldgcount")
                            if val is None and fallback_ea_feat is not None:
                                val = _extract_feature_attribute(fallback_ea_feat, ea_name_to_idx, "bldgcount")
                        elif ea_field_name == "hhcount":
                            val = _extract_feature_attribute(feat, repl_name_to_idx, "hh_count")
                            if val is None and fallback_ea_feat is not None:
                                val = _extract_feature_attribute(fallback_ea_feat, ea_name_to_idx, "hh_count")
                        elif ea_field_name == "bldgcount":
                            val = _extract_feature_attribute(feat, repl_name_to_idx, "bldg_count")
                            if val is None and fallback_ea_feat is not None:
                                val = _extract_feature_attribute(fallback_ea_feat, ea_name_to_idx, "bldg_count")
                        elif ea_field_name == "new_ean":
                            val = _extract_feature_attribute(feat, repl_name_to_idx, "ean")
                            if val is None and fallback_ea_feat is not None:
                                val = _extract_feature_attribute(fallback_ea_feat, ea_name_to_idx, "new_ean")
                                if val is None:
                                    val = _extract_feature_attribute(fallback_ea_feat, ea_name_to_idx, "ean")
                        elif ea_field_name == "ea_type":
                            val = _ea_type_from_layer_name(layer.name())

                    if val is not None:
                        # Safe type conversion based on target field type
                        try:
                            if ea_field.type() in (QVariant.Double, getattr(QVariant, 'Float', QVariant.Double)):
                                val = float(val)
                            elif ea_field.type() in (
                                QVariant.Int, getattr(QVariant, 'LongLong', QVariant.Int),
                                getattr(QVariant, 'UInt', QVariant.Int), getattr(QVariant, 'ULongLong', QVariant.Int)
                            ):
                                val = int(round(float(val)))
                            elif ea_field.type() == QVariant.String:
                                val = str(val)
                        except (ValueError, TypeError):
                            pass
                        out_feat.setAttribute(ea_i, val)

                replacement_features.append(out_feat)
                total_count += 1


        # Build the combined union in one call — far more efficient than
        # iterative combine() when there are many replacement geometries.
        union_geom: Optional[QgsGeometry] = None
        if geom_list:
            union_geom = QgsGeometry.unaryUnion(geom_list)
            if union_geom and not union_geom.isNull():
                union_geom = _repair_geometry(union_geom)

        return replacement_features, union_geom, total_count

    def _replace_ea_geometries(
        self,
        combined_replacement: QgsGeometry,
        out_fields: Optional[QgsFields] = None,
        replacement_features: Optional[list] = None,
    ) -> tuple[list[QgsFeature], list[QgsFeature], int]:
        """Subtract combined_replacement from each EA feature geometry.

        Returns:
          (remaining_features, ghost_features, modified_count)

        ``remaining_features`` — EA features whose geometry survived the
          difference operation (partial or full overlap removed).
        ``ghost_features`` — EA features that were *fully* consumed by a
          replacement polygon.  Their geometry is set to NULL so they are
          **not** added to the spatial output layer, but they are forwarded
          to the Excel exporter as reference rows tagged ``ea_type = MERGED``
          with ``eacount = NULL`` so subtotals remain correct.
        """
        if out_fields is None:
            out_fields = self._build_output_fields()
        # Build spatial index on EA layer for efficient candidate selection
        ea_index = QgsSpatialIndex(self.ea_layer.getFeatures())
        combined_bbox = combined_replacement.boundingBox()
        candidate_id_set = set(ea_index.intersects(combined_bbox))

        ea_fields = self.ea_layer.fields()
        ea_name_to_idx = {
            ea_fields.at(i).name().lower(): i
            for i in range(ea_fields.count())
        }

        # Field indices in the output schema needed for ghost-row annotation
        out_name_to_idx = {
            out_fields.at(i).name().lower(): i
            for i in range(out_fields.count())
        }
        _ghost_ea_type_idx  = out_name_to_idx.get("ea_type",  -1)
        _ghost_eacount_idx  = out_name_to_idx.get("eacount",  -1)
        _ghost_remarks_idx  = out_name_to_idx.get("remarks",  -1)

        remaining_features: list[QgsFeature] = []
        ghost_features: list[QgsFeature] = []
        modified_count = 0

        for feat in self.ea_layer.getFeatures():
            fid = feat.id()
            ea_geom = QgsGeometry(feat.geometry())
            if ea_geom is None or ea_geom.isNull() or ea_geom.isEmpty():
                continue

            ea_geom = _repair_geometry(ea_geom)
            geom_to_keep = None
            fully_consumed = False

            if fid in candidate_id_set:
                # Spatial index says bounding boxes overlap — confirm with exact
                # geometric intersection before computing the difference.
                if ea_geom.intersects(combined_replacement):
                    orig_area = ea_geom.area()
                    inter = ea_geom.intersection(combined_replacement)
                    inter_area = inter.area() if (inter and not inter.isNull() and not inter.isEmpty()) else 0.0

                    # Relative overlap ratio: works identically for projected (meters) and geographic (degrees).
                    # If >= 99.9% of the original EA is covered by the replacement polygon, treat as fully consumed.
                    if orig_area > 0 and (inter_area / orig_area) >= 0.999:
                        modified_count += 1
                        fully_consumed = True
                    else:
                        remaining = ea_geom.difference(combined_replacement)
                        if remaining is None or remaining.isNull() or remaining.isEmpty():
                            modified_count += 1
                            fully_consumed = True
                        else:
                            remaining = _repair_geometry(remaining)
                            if remaining is None or remaining.isNull() or remaining.isEmpty():
                                modified_count += 1
                                fully_consumed = True
                            else:
                                rem_area = remaining.area()
                                # Clean up tiny sliver fragments (< 0.1% of original area)
                                if orig_area > 0 and (rem_area / orig_area) < 0.001:
                                    modified_count += 1
                                    fully_consumed = True
                                else:
                                    if abs(orig_area - rem_area) > (orig_area * 1e-6):
                                        modified_count += 1
                                    geom_to_keep = remaining
                else:
                    # Bounding-box overlap but no actual geometric intersection
                    geom_to_keep = ea_geom
            else:
                # No spatial-index overlap — keep as-is, no geometry test needed
                geom_to_keep = ea_geom

            # ── Build output feature (remaining or ghost) ──────────────────
            out_feat = QgsFeature(out_fields)
            if fully_consumed:
                # Geometry intentionally left NULL — ghost row for Excel only
                out_feat.setGeometry(QgsGeometry())
            else:
                out_feat.setGeometry(geom_to_keep)

            for ea_i in range(out_fields.count()):
                ea_field = out_fields.at(ea_i)
                ea_field_name = ea_field.name().lower()
                val = _extract_feature_attribute(feat, ea_name_to_idx, ea_field_name)

                if val is None:
                    if ea_field_name == "hh_count":
                        val = _extract_feature_attribute(feat, ea_name_to_idx, "hhcount")
                    elif ea_field_name == "bldg_count":
                        val = _extract_feature_attribute(feat, ea_name_to_idx, "bldgcount")
                    elif ea_field_name == "hhcount":
                        val = _extract_feature_attribute(feat, ea_name_to_idx, "hh_count")
                    elif ea_field_name == "bldgcount":
                        val = _extract_feature_attribute(feat, ea_name_to_idx, "bldg_count")
                    elif ea_field_name == "new_ean":
                        val = _extract_feature_attribute(feat, ea_name_to_idx, "ean")
                    elif ea_field_name == "ea_type":
                        val = "RETAINED"

                if val is not None:
                    try:
                        if ea_field.type() in (QVariant.Double, getattr(QVariant, 'Float', QVariant.Double)):
                            val = float(val)
                        elif ea_field.type() in (
                            QVariant.Int, getattr(QVariant, 'LongLong', QVariant.Int),
                            getattr(QVariant, 'UInt', QVariant.Int), getattr(QVariant, 'ULongLong', QVariant.Int)
                        ):
                            val = int(round(float(val)))
                        elif ea_field.type() == QVariant.String:
                            val = str(val)
                    except (ValueError, TypeError):
                        pass
                    out_feat.setAttribute(ea_i, val)

            if fully_consumed:
                # Override ea_type → MERGED and clear eacount (reference row only)
                if _ghost_ea_type_idx != -1:
                    out_feat.setAttribute(_ghost_ea_type_idx, "MERGED")
                if _ghost_eacount_idx != -1:
                    out_feat.setAttribute(_ghost_eacount_idx, None)

                # Annotate remarks with the absorbing replacement's new_ean
                absorbing_repl_feat, absorbing_new_ean = self._absorbing_replacement_feat_and_new_ean(
                    ea_geom, replacement_features or []
                )
                consumed_ea_code = str(_extract_feature_attribute(feat, ea_name_to_idx, "ean") or "")

                existing_remarks = ""
                if _ghost_remarks_idx != -1:
                    raw_rmk = out_feat.attribute(_ghost_remarks_idx)
                    from qgis.core import NULL
                    if raw_rmk is not None and raw_rmk != NULL:
                        s_rmk = str(raw_rmk).strip()
                        if s_rmk.upper() not in ("", "NULL", "NONE", "FALSE", "0", "F"):
                            existing_remarks = s_rmk

                if absorbing_new_ean:
                    merged_note = f"Merged to EA {absorbing_new_ean}"
                else:
                    merged_note = "Merged EA"

                if existing_remarks and "merge" not in existing_remarks.lower():
                    new_remarks = f"{merged_note}; {existing_remarks}"
                elif not existing_remarks:
                    new_remarks = merged_note
                else:
                    new_remarks = existing_remarks  # already annotated

                if _ghost_remarks_idx != -1:
                    out_feat.setAttribute(_ghost_remarks_idx, new_remarks)

                # Also annotate the absorbing replacement feature's remarks with the absorbed EA
                if absorbing_repl_feat is not None and consumed_ea_code and _ghost_remarks_idx != -1:
                    repl_rmk = str(absorbing_repl_feat.attribute(_ghost_remarks_idx) or "").strip()
                    if repl_rmk.upper() in ("", "NULL", "NONE", "FALSE", "0", "F"):
                        repl_rmk = ""
                    if consumed_ea_code != absorbing_new_ean and consumed_ea_code not in repl_rmk:
                        if not repl_rmk:
                            absorbing_repl_feat.setAttribute(_ghost_remarks_idx, f"Merged with EA {consumed_ea_code}")
                        elif "merged with" in repl_rmk.lower():
                            absorbing_repl_feat.setAttribute(_ghost_remarks_idx, f"{repl_rmk}, {consumed_ea_code}")
                        else:
                            absorbing_repl_feat.setAttribute(_ghost_remarks_idx, f"Merged with EA {consumed_ea_code}; {repl_rmk}")

                ghost_features.append(out_feat)
            else:
                remaining_features.append(out_feat)

        return remaining_features, ghost_features, modified_count

    def _absorbing_replacement_feat_and_new_ean(
        self,
        consumed_geom: QgsGeometry,
        replacement_features: list,
    ) -> tuple[Optional[QgsFeature], Optional[str]]:
        """Find the replacement feature and its new_ean that most overlaps consumed_geom.

        Iterates through already-built replacement QgsFeature objects, computes
        intersection area, and returns (best_repl_feat, best_new_ean).
        """
        if not replacement_features or consumed_geom is None or consumed_geom.isNull():
            return None, None

        from qgis.core import NULL

        best_area = 0.0
        best_repl_feat: Optional[QgsFeature] = None
        best_new_ean: Optional[str] = None

        for repl_feat in replacement_features:
            repl_geom = repl_feat.geometry()
            if repl_geom is None or repl_geom.isNull() or repl_geom.isEmpty():
                continue
            inter = consumed_geom.intersection(repl_geom)
            if inter is None or inter.isNull() or inter.isEmpty():
                continue
            area = inter.area()
            if area > best_area:
                best_area = area
                best_repl_feat = repl_feat
                # Read new_ean from the replacement feature's attributes
                repl_fields = repl_feat.fields()
                if repl_fields is not None:
                    repl_name_to_idx = {
                        repl_fields.at(i).name().lower(): i
                        for i in range(repl_fields.count())
                    }
                    raw = _extract_feature_attribute(
                        repl_feat, repl_name_to_idx, "new_ean"
                    )
                    if raw is None:
                        raw = _extract_feature_attribute(
                            repl_feat, repl_name_to_idx, "ean"
                        )
                    if raw is None:
                        for fname in ("new_eacode", "new_ea", "ea_no", "eano", "ea"):
                            raw = _extract_feature_attribute(repl_feat, repl_name_to_idx, fname)
                            if raw is not None:
                                break

                    if raw is not None and raw != NULL:
                        s = str(raw).strip()
                        if s not in ("", "NULL", "None"):
                            if s.isdigit() and len(s) < 6:
                                s = s.zfill(6)
                            best_new_ean = s

        return best_repl_feat, best_new_ean

    def _populate_ea_counts(
        self,
        features: list[QgsFeature],
        out_fields: QgsFields,
    ) -> None:
        """Populate the EACount field based on the 8-digit geocode and new_ean.

        Assigns 1 on the first occurrence of each unique (8-digit geocode, new_ean)
        pair, and leaves EACount empty (None / NULL) on duplicate rows/fragments
        to prevent double-counting due to delineation.
        """
        # Find the EACount field index in out_fields
        ea_count_idx = -1
        for i in range(out_fields.count()):
            if out_fields.at(i).name().lower() == "eacount":
                ea_count_idx = i
                break

        if ea_count_idx == -1:
            return

        out_name_to_idx = {
            out_fields.at(i).name().lower(): i
            for i in range(out_fields.count())
        }

        seen_ea_keys = set()

        for idx, feat in enumerate(features):
            raw_geo = _extract_feature_attribute(feat, out_name_to_idx, "geocode")
            if not raw_geo:
                raw_geo = _extract_feature_attribute(feat, out_name_to_idx, "barangay")
            if not raw_geo:
                raw_geo = _extract_feature_attribute(feat, out_name_to_idx, "code")
            if not raw_geo:
                raw_geo = self._geo_code or ""

            digits = re.sub(r"\D", "", str(raw_geo).strip())
            if len(digits) >= 8:
                bgy_code = digits[:8]
            elif digits:
                bgy_code = digits
            else:
                bgy_code = str(raw_geo).strip() or "UNKNOWN"

            ean_val = _extract_feature_attribute(feat, out_name_to_idx, "new_ean")
            if ean_val is None or str(ean_val).strip() in ("", "NULL", "None"):
                ean_val = _extract_feature_attribute(feat, out_name_to_idx, "ean")
            if ean_val is None or str(ean_val).strip() in ("", "NULL", "None"):
                ean_val = f"feat_{idx}"

            ean_str = str(ean_val).strip()
            ea_key = (bgy_code, ean_str)

            if ea_key not in seen_ea_keys:
                seen_ea_keys.add(ea_key)
                feat.setAttribute(ea_count_idx, 1)
            else:
                feat.setAttribute(ea_count_idx, None)

    def _create_output_layer(
        self,
        features: list[QgsFeature],
        out_fields: Optional[QgsFields] = None,
    ) -> Optional[QgsVectorLayer]:
        """Build the output in-memory polygon layer from the collected features.

        Uses out_fields containing Previous EA Layer attributes plus all count fields.
        """
        if out_fields is None:
            out_fields = self._build_output_fields()
        crs_auth = self.ea_layer.crs().authid()
        is_multi = False
        try:
            if hasattr(QgsWkbTypes, 'isMultiType') and callable(getattr(QgsWkbTypes, 'isMultiType')):
                is_multi = QgsWkbTypes.isMultiType(self.ea_layer.wkbType())
            else:
                is_multi = "multi" in str(self.ea_layer.wkbType()).lower()
        except Exception:
            is_multi = False

        wkb_type = "MultiPolygon" if is_multi else "Polygon"
        uri = f"{wkb_type}?crs={crs_auth}"
        layer = QgsVectorLayer(uri, self._output_layer_name, "memory")
        if not layer.isValid():
            self._log("[ERROR] Failed to create memory layer.")
            return None

        provider = layer.dataProvider()

        # Set field schema from out_fields
        fields_list = [out_fields.at(i) for i in range(out_fields.count())]
        provider.addAttributes(fields_list)
        layer.updateFields()

        # Ensure all output features have unique sequential FIDs
        fid_idx = layer.fields().indexOf("fid")
        for idx, feat in enumerate(features, start=1):
            if feat.fields() is None or feat.fields().count() != layer.fields().count():
                attrs = list(feat.attributes()) if feat.attributes() else []
                if len(attrs) < layer.fields().count():
                    attrs.extend([None] * (layer.fields().count() - len(attrs)))
                feat.setFields(layer.fields())
                feat.setAttributes(attrs)
            if fid_idx != -1:
                feat.setAttribute(fid_idx, idx)
            feat.setId(idx)

        # Add features in bulk
        provider.addFeatures(features)
        layer.updateExtents()

        return layer

    def _validate_output(self, layer: QgsVectorLayer) -> bool:
        """Perform basic validation on the output layer."""
        if layer is None or not layer.isValid():
            self._log("[WARNING] Output layer is invalid.")
            return False
        if layer.featureCount() == 0:
            self._log("[WARNING] Output layer contains no features.")
            return False
        self._log(
            f"[INFO] Output layer validation passed: "
            f"{layer.featureCount()} features."
        )
        return True

    def _export_layer_to_gpkg(
        self, layer: QgsVectorLayer, file_path: str, layer_name: str
    ) -> bool:
        """Export a vector layer to a permanent GeoPackage (.gpkg) file on disk."""
        try:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # Method 1: QGIS Processing native:savefeatures (fast C++ engine)
            try:
                import processing
                params = {
                    "INPUT": layer,
                    "OUTPUT": file_path,
                    "LAYER_NAME": layer_name,
                }
                res = processing.run("native:savefeatures", params)
                if res and os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                    self._log(f"[INFO] Layer successfully written to GeoPackage via QGIS Processing: {file_path}")
                    return True
            except Exception as pe:
                self._log(f"[DEBUG] QGIS Processing native:savefeatures fallback: {pe}")

            from qgis.core import (
                QgsCoordinateTransformContext,
                QgsVectorFileWriter,
            )

            # Method 2: SaveVectorOptions with writeAsVectorFormatV3 / V2
            save_options = QgsVectorFileWriter.SaveVectorOptions()
            save_options.driverName = "GPKG"
            save_options.layerName = layer_name
            save_options.fileEncoding = "UTF-8"
            if os.path.exists(file_path):
                save_options.actionOnExistingFile = QgsVectorFileWriter.CreateOrOverwriteLayer
            else:
                save_options.actionOnExistingFile = QgsVectorFileWriter.CreateOrOverwriteFile

            ctx = (
                QgsProject.instance().transformContext()
                if (QgsProject.instance() and hasattr(QgsProject.instance(), 'transformContext'))
                else QgsCoordinateTransformContext()
            )

            if hasattr(QgsVectorFileWriter, 'writeAsVectorFormatV3'):
                res = QgsVectorFileWriter.writeAsVectorFormatV3(layer, file_path, ctx, save_options)
                if res[0] == QgsVectorFileWriter.NoError:
                    return True
                self._log(f"[DEBUG] writeAsVectorFormatV3 code={res[0]}: {res[1] if len(res) > 1 else ''}")

            if hasattr(QgsVectorFileWriter, 'writeAsVectorFormatV2'):
                res = QgsVectorFileWriter.writeAsVectorFormatV2(layer, file_path, ctx, save_options)
                if res[0] == QgsVectorFileWriter.NoError:
                    return True
                self._log(f"[DEBUG] writeAsVectorFormatV2 code={res[0]}: {res[1] if len(res) > 1 else ''}")

            # Method 3: Legacy writeAsVectorFormat
            if hasattr(QgsVectorFileWriter, 'writeAsVectorFormat'):
                res = QgsVectorFileWriter.writeAsVectorFormat(layer, file_path, "UTF-8", layer.crs(), "GPKG")
                if res == QgsVectorFileWriter.NoError:
                    return True
                self._log(f"[DEBUG] writeAsVectorFormat code={res}")

            # Method 4: Direct writer feature-by-feature
            writer = QgsVectorFileWriter(file_path, "UTF-8", layer.fields(), layer.wkbType(), layer.crs(), "GPKG")
            if writer.hasError() == QgsVectorFileWriter.NoError:
                for feat in layer.getFeatures():
                    writer.addFeature(feat)
                del writer
                return True
            else:
                self._log(f"[DEBUG] Direct QgsVectorFileWriter error={writer.errorMessage()}")

        except Exception as e:
            self._log(f"[ERROR] Exception during GeoPackage export: {e}")
            return False

        return os.path.exists(file_path) and os.path.getsize(file_path) > 0

    def _export_attribute_table_to_excel(
        self,
        layer: QgsVectorLayer,
        path: str,
        ghost_features: Optional[list] = None,
    ) -> bool:
        """Export the layer attribute table to Excel.

        Columns = layer fields (same order, same names, no extra columns).
        Rows    = all features (no geometry) followed by ghost rows.
        Sheet   = EA2026.

        ghost_features : list[QgsFeature], optional
            Fully-consumed previous EA features (NULL geometry) that were
            excluded from the spatial output layer but should appear in the
            Excel file as reference rows tagged ea_type=MERGED.
        """
        try:
            import openpyxl
            from openpyxl.writer.excel import save_workbook  # noqa: F401
        except ImportError:
            self._log(
                "[ERROR] openpyxl is not installed. "
                "Cannot generate Excel output."
            )
            return False

        try:
            # Use write_only=True to avoid the Windows fatal access violation
            # that occurs when openpyxl.Workbook() calls _setup_styles() →
            # copy(DEFAULT_FONT) → to_tree() → lxml Element(), which clashes
            # with QGIS's own libxml2 critical-section lock on the main thread.
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet(title=_EXCEL_SHEET_NAME)

            fields = layer.fields()
            field_names = [fields.at(i).name() for i in range(fields.count())]

            # Header row
            ws.append(field_names)

            from qgis.core import NULL

            def _row_values(feat, flds):
                row = []
                for i in range(flds.count()):
                    val = feat.attribute(i)
                    row.append(None if (val == NULL or val is None) else val)
                return row

            # Data rows — spatial features
            active_keys = set()
            for feat in layer.getFeatures(
                QgsFeatureRequest().setFlags(QgsFeatureRequest.NoGeometry)
            ):
                ws.append(_row_values(feat, fields))
                g_code = str(feat.attribute("geocode") if "geocode" in field_names else (feat.attribute("GEOCODE") if "GEOCODE" in field_names else ""))[:8]
                e_code = str(feat.attribute("ean") if "ean" in field_names else (feat.attribute("EA_NO") if "EA_NO" in field_names else (feat.attribute("ea") if "ea" in field_names else "")))
                if e_code:
                    active_keys.add((g_code, e_code))

            # Ghost rows — fully-consumed previous EAs (Excel-only reference rows)
            for ghost_feat in (ghost_features or []):
                ghost_fields = ghost_feat.fields()
                if ghost_fields is None or ghost_fields.count() == 0:
                    continue
                # Align ghost feature attributes to the layer's field order
                ghost_name_to_idx = {
                    ghost_fields.at(i).name().lower(): i
                    for i in range(ghost_fields.count())
                }

                g_gcode = str(ghost_feat.attribute(ghost_name_to_idx.get("geocode", -1)) if "geocode" in ghost_name_to_idx else "")[:8]
                g_ean = str(ghost_feat.attribute(ghost_name_to_idx.get("ean", -1)) if "ean" in ghost_name_to_idx else (ghost_feat.attribute(ghost_name_to_idx.get("ea_no", -1)) if "ea_no" in ghost_name_to_idx else (ghost_feat.attribute(ghost_name_to_idx.get("ea", -1)) if "ea" in ghost_name_to_idx else "")))
                if g_ean and (g_gcode, g_ean) in active_keys:
                    # Already represented by an active spatial row; do not duplicate
                    continue

                row = []
                for fname in field_names:
                    idx = ghost_name_to_idx.get(fname.lower(), -1)
                    if idx == -1:
                        row.append(None)
                    else:
                        val = ghost_feat.attribute(idx)
                        row.append(None if (val == NULL or val is None) else val)
                ws.append(row)

            wb.save(path)
            return True

        except Exception as exc:
            self._log(f"[ERROR] Excel export failed: {exc}")
            return False

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _log(self, msg: str) -> None:
        """Send a log message to the feedback callback and store it."""
        self._result.log_lines.append(msg)
        self._feedback(msg)

    def _fail(self, msg: str) -> None:
        """Record a failure: log the error and mark the result as failed."""
        self._result.success = False
        self._result.error_message = msg
        self._result.summary.overall_status = "ERROR"
        self._log(f"[ERROR] {msg}")
